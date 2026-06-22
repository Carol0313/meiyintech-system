"""
商户平台视图
包含：首页、会员管理、工厂管理、商品规格、订单管理、拼版工具、权限与子账号
"""

import json
import logging
import os
import uuid
from decimal import Decimal
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db import transaction, models
from django.utils import timezone
from django.http import HttpResponse
from django.core.files.storage import default_storage
from apps.accounts.models import User, CustomerProfile, Merchant, StaffProfile, Role

logger = logging.getLogger(__name__)
from apps.orders.models import Order, OrderItem, CommunicationLog, OrderStatusLog, PlateLayout, PlateBatch, PlateBatchItem, ProductionPhoto, Statement, DeliveryExtension, OrderComplaint
from apps.products.models import ProductSpec, CustomSpecRequest
from .models import Factory, FactoryEquipmentStatus, FactoryInventory
from utils.plate_layout import calculate_plate_layout, calculate_plate_layout_rectpack, auto_generate_plate_layout_for_order
from .permissions import staff_permission


def merchant_required(view_func):
    """商家管理员权限装饰器"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, '请先登录商家账号')
            return redirect('login')
        if request.user.user_type not in ('merchant_admin', 'merchant_staff'):
            return render(request, 'common/account_unauthorized.html', {
                'message': '请先登录商家账号',
                'can_logout': True,
            })
        # 账号必须激活且已审核
        if not request.user.is_active or not request.user.is_approved:
            return render(request, 'common/account_unauthorized.html', {
                'message': '账号未启用或未审核通过，请联系管理员处理。',
                'can_logout': True,
            })
        # 商家管理员检查 managed_merchant
        merchant = None
        if request.user.user_type == 'merchant_admin':
            merchant = getattr(request.user, 'managed_merchant', None)
            if not merchant:
                return render(request, 'common/account_unauthorized.html', {
                    'message': '您没有关联的商家，请联系平台管理员处理。',
                    'can_logout': True,
                })
        else:
            # 商家员工检查 staff_profile
            profile = getattr(request.user, 'staff_profile', None)
            if not profile or not profile.is_active:
                return render(request, 'common/account_unauthorized.html', {
                    'message': '您的员工账号未启用，请联系商家管理员处理。',
                    'can_logout': True,
                })
            merchant = profile.merchant
        # 商家必须处于已通过状态（冻结/拒绝/待审核均不可访问）
        if not merchant or merchant.status != 'approved':
            return render(request, 'common/account_unauthorized.html', {
                'message': '商家未通过审核或已被冻结，暂无法访问后台。',
                'can_logout': True,
            })
        return view_func(request, *args, **kwargs)
    return wrapper


def merchant_admin_required(view_func):
    """仅商家总管理员（须先通过 merchant_required 校验）"""
    @merchant_required
    def wrapper(request, *args, **kwargs):
        if request.user.user_type != 'merchant_admin':
            messages.error(request, '仅商家管理员可操作')
            return redirect('merchant_dashboard')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def get_merchant(request):
    """获取当前用户关联的商家"""
    if request.user.user_type == 'merchant_admin':
        return getattr(request.user, 'managed_merchant', None)
    profile = getattr(request.user, 'staff_profile', None)
    return profile.merchant if profile else None



# ==================== 商家首页 ====================

# 制版行业每日一句小知识
DAILY_TIPS = [
    "腐蚀版的最细线条建议不小于0.12mm，否则晒版时容易断线。",
    "烫金版使用阴片，压纹版使用阳片，下单时务必确认清楚。",
    "镁版密度仅为1.74g/cm³，比铜版轻70%，适合大面积烫金。",
    "雕刻版的拼版间距通常比腐蚀版大5mm，以确保CNC雕刻时不会互相干扰。",
    "树脂版水洗工艺比倒模工艺更环保，但耐印率略低约15%。",
    "菲林输出前必须确认文件已转曲，否则字体替换会导致跑位。",
    "浮雕版（多层次）的最小间隙建议不小于0.10mm，避免雕刻时崩边。",
    "加急订单建议提前与工厂确认排期，避免与其他订单冲突导致延误。",
    "制版文件推荐使用AI 3.0格式保存，兼容性最佳且不易丢数据。",
    "红框尺寸识别可以帮助设计师更准确地计算实际制版面积，减少误差。",
    "不同厚度的版材拼版时，间距应增加至20mm以上，防止压印时互相影响。",
    "平雕版与浮雕版的价格差异主要取决于雕刻深度、层次数量和材质。",
    "铜版导热性优于镁版，适合高速烫金机长时间连续作业。",
    "订单文件上传前建议先进行PDF预检，确保单色K100%，避免四色黑。",
    "信用额度支付可以帮助大客户简化流程，月底统一对账更方便。",
    "6.35mm厚度的镁版常用于深压纹，压印深度可达2-3mm。",
    "菲林对位时需要保留角线，裁切后角线宽度不应小于0.5mm。",
    "腐蚀版在显影后要仔细检查是否有砂眼，可用红墨水补涂修复。",
    "雕刻版的刀具直径最小可达0.1mm，但过细容易崩刀，建议0.2mm以上。",
    "版材存放应避免潮湿环境，镁版在湿度>70%时容易氧化发黑。",
    "补版单无需重新支付，系统会自动关联原订单的拼版数据。",
    "烫金温度建议控制在140-180℃之间，过高会导致烫金纸粘版。",
    "UV菲林与普通对位菲林的区别在于UV菲林需要更高的线条对比度。",
    "激凸版的内角线距离图案不足10mm时，建议保留5mm安全间距。",
    "不锈钢版耐腐蚀性最佳，适合化妆品、食品等高标准包装烫金。",
]


def get_daily_tip():
    """根据日期获取每日一句（每天固定一句，全年循环）"""
    from datetime import date
    day_index = date.today().toordinal()
    return DAILY_TIPS[day_index % len(DAILY_TIPS)]


@login_required
@merchant_required
def merchant_dashboard(request):
    """商家首页"""
    merchant = get_merchant(request)
    today = timezone.now().date()
    pending_orders = merchant.orders.filter(status='pending_confirm').count()
    in_production = merchant.orders.filter(status='in_production').count()
    recent_orders = merchant.orders.order_by('-created_at')[:5]
    customer_count = merchant.customers.filter(registration_status='approved').count()
    
    # 投诉统计
    merchant_complaints = OrderComplaint.objects.filter(order__merchant=merchant)
    complaint_pending = merchant_complaints.filter(status='pending').count()
    complaint_processing = merchant_complaints.filter(status='processing').count()
    complaint_total = merchant_complaints.count()
    
    ctx = {
        'merchant': merchant,
        'pending_orders': pending_orders,
        'in_production': in_production,
        'customer_count': customer_count,
        'recent_orders': recent_orders,
        'daily_tip': get_daily_tip(),
        'complaint_pending': complaint_pending,
        'complaint_processing': complaint_processing,
        'complaint_total': complaint_total,
    }
    return render(request, 'merchant/dashboard.html', ctx)


# ==================== 数据分析中心 ====================

@login_required
@merchant_required
def merchant_analytics(request):
    """商户数据分析中心 - 经营数据可视化"""
    from django.db.models import Sum, Count, Avg, Q, F
    from datetime import timedelta
    from django.utils import timezone
    from apps.products.models import ProductSpec

    merchant = get_merchant(request)

    # ---- 时间范围参数 ----
    range_key = request.GET.get('range', 'month')
    now = timezone.now()
    today = now.date()

    if range_key == 'today':
        start_date = today
        end_date = today
    elif range_key == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif range_key == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif range_key == 'quarter':
        quarter_start_month = (today.month - 1) // 3 * 3 + 1
        start_date = today.replace(month=quarter_start_month, day=1)
        end_date = today
    elif range_key == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    else:
        range_key = 'month'
        start_date = today.replace(day=1)
        end_date = today

    # 基础查询集（已提交的订单）
    base_orders = merchant.orders.filter(is_submitted=True)
    date_orders = base_orders.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

    # ---- KPI 统计 ----
    total_orders = date_orders.count()
    total_amount = date_orders.aggregate(s=Sum('total_amount'))['s'] or 0
    total_customers = date_orders.values('customer').distinct().count()

    # 总版数 + 总面积
    date_items = OrderItem.objects.filter(order__in=date_orders)
    total_plates = date_items.aggregate(s=Sum('quantity'))['s'] or 0
    total_area = date_items.aggregate(s=Sum('area'))['s'] or 0

    avg_order_value = total_amount / total_orders if total_orders > 0 else 0

    remake_count = date_orders.filter(order_type='remake').count()
    remake_rate = round(remake_count / total_orders * 100, 1) if total_orders > 0 else 0

    # ---- 订单趋势（按天）----
    from django.db.models.functions import TruncDate
    trend_data = list(date_orders.annotate(
        day=TruncDate('created_at')
    ).values('day').annotate(
        amount=Sum('total_amount'),
        count=Count('id')
    ).order_by('day'))

    trend_labels = [d['day'].strftime('%m-%d') for d in trend_data]
    trend_amounts = [float(d['amount'] or 0) for d in trend_data]
    trend_counts = [d['count'] for d in trend_data]

    # ---- 产品类型分布 ----
    product_stats = list(date_items.values('product_name').annotate(
        order_count=Count('order', distinct=True),
        total_qty=Sum('quantity')
    ).order_by('-total_qty'))
    product_name_map = {p[0]: p[1] for p in ProductSpec.PRODUCT_NAME_CHOICES}
    for s in product_stats:
        s['label'] = product_name_map.get(s['product_name'], s['product_name'])

    # ---- 材质分布 ----
    material_stats = list(date_items.values('material').annotate(
        total_area=Sum('area'),
        total_qty=Sum('quantity')
    ).order_by('-total_area'))
    material_name_map = dict(OrderItem.MATERIAL_CHOICES)
    for s in material_stats:
        s['label'] = material_name_map.get(s['material'], s['material'])

    # ---- 客户价值排行 Top 10 ----
    customer_stats = list(date_orders.values('customer__phone').annotate(
        total_amount=Sum('total_amount'),
        order_count=Count('id')
    ).order_by('-total_amount')[:10])
    for c in customer_stats:
        phone = c['customer__phone'] or ''
        c['phone_display'] = f"{phone[:3]}****{phone[-4:]}" if len(phone) >= 7 else phone

    # ---- 工厂产能对比 ----
    factory_stats = list(date_orders.filter(factory__isnull=False).values(
        'factory__name'
    ).annotate(
        order_count=Count('id'),
        total_amount=Sum('total_amount')
    ).order_by('-order_count'))
    # 补充面积数据
    factory_area_qs = list(date_items.filter(order__factory__isnull=False).values(
        'order__factory__name'
    ).annotate(total_area=Sum('area')))
    factory_area_stats = {
        fa.get('order__factory__name', ''): fa.get('total_area', 0)
        for fa in factory_area_qs
    }
    for f in factory_stats:
        f['total_area'] = factory_area_stats.get(f.get('factory__name', ''), 0)

    # ---- 订单状态分布 ----
    status_distribution = []
    status_flow = [
        ('pending_confirm', '待确认'),
        ('design_confirmed', '设计确认'),
        ('paid', '已付款'),
        ('in_production', '生产中'),
        ('shipped', '已发货'),
        ('received', '已收货'),
    ]
    for code, label in status_flow:
        count = date_orders.filter(status=code).count()
        status_distribution.append({'code': code, 'label': label, 'count': count})

    # ---- SLA 时效统计 ----
    sla_stats = {
        'cs_avg_minutes': None,
        'factory_avg_minutes': None,
        'cs_overdue_count': 0,
        'factory_overdue_count': 0,
    }

    # 客服平均处理时长
    cs_orders = date_orders.filter(
        file_uploaded_at__isnull=False,
        customer_service_processed_at__isnull=False
    )
    cs_times = []
    for o in cs_orders:
        delta = (o.customer_service_processed_at - o.file_uploaded_at).total_seconds() / 60
        if delta > 0:
            cs_times.append(delta)
    if cs_times:
        sla_stats['cs_avg_minutes'] = round(sum(cs_times) / len(cs_times), 1)

    # 工厂平均下载时长
    factory_times = []
    factory_orders = date_orders.filter(
        factory_notified_at__isnull=False,
        factory_downloaded_at__isnull=False
    )
    for o in factory_orders:
        delta = (o.factory_downloaded_at - o.factory_notified_at).total_seconds() / 60
        if delta > 0:
            factory_times.append(delta)
    if factory_times:
        sla_stats['factory_avg_minutes'] = round(sum(factory_times) / len(factory_times), 1)

    # 超时统计
    cs_pending = date_orders.filter(
        file_uploaded_at__isnull=False,
        customer_service_processed_at__isnull=False
    )
    for o in cs_pending:
        elapsed = (o.customer_service_processed_at - o.file_uploaded_at).total_seconds() / 60
        if elapsed > 30:
            sla_stats['cs_overdue_count'] += 1

    factory_pending = date_orders.filter(
        factory_notified_at__isnull=False,
        factory_downloaded_at__isnull=False
    )
    for o in factory_pending:
        elapsed = (o.factory_downloaded_at - o.factory_notified_at).total_seconds() / 60
        if elapsed > 30:
            sla_stats['factory_overdue_count'] += 1

    ctx = {
        'range_key': range_key,
        'start_date': start_date,
        'end_date': end_date,
        'kpi': {
            'total_orders': total_orders,
            'total_amount': total_amount,
            'total_plates': total_plates,
            'total_area': total_area,
            'avg_order_value': avg_order_value,
            'total_customers': total_customers,
            'remake_count': remake_count,
            'remake_rate': remake_rate,
        },
        'trend_labels': trend_labels,
        'trend_amounts': trend_amounts,
        'trend_counts': trend_counts,
        'product_stats': product_stats,
        'material_stats': material_stats,
        'customer_stats': customer_stats,
        'factory_stats': factory_stats,
        'status_distribution': status_distribution,
        'sla_stats': sla_stats,
    }
    return render(request, 'merchant/analytics.html', ctx)


# ==================== 账号设置 ====================

@login_required
@merchant_required
def merchant_settings(request):
    """商家账号设置"""
    merchant = get_merchant(request)
    if request.method == 'POST':
        merchant.name = request.POST.get('name', merchant.name)
        merchant.address = request.POST.get('address', merchant.address)
        merchant.contact_phone = request.POST.get('contact_phone', merchant.contact_phone)
        merchant.customer_service_wechat = request.POST.get('customer_service_wechat', merchant.customer_service_wechat)
        regions = request.POST.get('service_regions', '')
        # 尝试解析JSON格式的新数据，兼容旧格式
        try:
            if regions and regions.startswith('['):
                import json
                region_data = json.loads(regions)
                # 存储为JSON字符串
                merchant.service_regions = regions
            else:
                merchant.service_regions = regions
        except:
            merchant.service_regions = regions
        merchant.save()
        messages.success(request, '商家信息已更新')
        return redirect('merchant_settings')
    return render(request, 'merchant/settings.html', {'merchant': merchant})


# ==================== 会员管理 ====================

@login_required
@merchant_required
def member_list(request):
    """会员列表与审核"""
    merchant = get_merchant(request)
    status_filter = request.GET.get('status')
    members = merchant.customers.all()
    if status_filter:
        members = members.filter(registration_status=status_filter)
    members = members.order_by('-created_at')
    
    # 为每个客户预计算未对账金额和订单数
    from django.db.models import Sum, Count, Q
    for m in members:
        unsettled_stats = Order.objects.filter(
            customer=m.user,
            merchant=merchant,
            status='received',
            is_settled=False,
            statement__isnull=True,
        ).aggregate(total=Sum('total_amount'), count=Count('id'))
        m.unsettled_amount = unsettled_stats['total'] or 0
        m.unsettled_count = unsettled_stats['count'] or 0
    
    return render(request, 'merchant/member_list.html', {
        'members': members,
        'status_filter': status_filter,
        'merchant': merchant,
    })


@login_required
@merchant_required
def member_pricing(request, profile_id):
    """客户报价单：商户为待审核客户填写各规格单价，系统根据城市档位核对差异"""
    merchant = get_merchant(request)
    profile = get_object_or_404(CustomerProfile, pk=profile_id, merchant=merchant)

    from utils.pricing_tiers import (
        get_product_category, is_etching_product,
        get_etching_price, get_carving_price, TIER_PRICES
    )

    if request.method == 'POST':
        import json
        custom_prices = {}
        for key, value in request.POST.items():
            if key.startswith('price_'):
                spec_key = key[6:]  # 去掉 "price_" 前缀
                try:
                    val = Decimal(value)
                    if val >= 0:
                        custom_prices[spec_key] = str(val.quantize(Decimal('0.01')))
                except:
                    pass
        profile.custom_prices = json.dumps(custom_prices, ensure_ascii=False)
        profile.credit_limit = Decimal(request.POST.get('credit_limit', '10000'))
        profile.registration_status = 'approved'
        profile.save()
        profile.user.is_approved = True
        profile.user.save(update_fields=['is_approved'])
        messages.success(request, f'已保存报价并审核通过 {profile.real_name or profile.user.phone}')
        return redirect('member_list')

    # GET: 构建报价单数据
    specs = ProductSpec.objects.filter(is_platform_preset=True).order_by('product_name', 'material', 'thickness')
    groups = {}
    for s in specs:
        category = get_product_category(s.product_name)
        if category not in groups:
            groups[category] = []
        key = f"{s.product_name}_{s.material}_{s.thickness}"
        if is_etching_product(s.product_name):
            ref_price = get_etching_price(profile.pricing_tier, s.thickness)
        else:
            ref_price = get_carving_price(s.product_name, s.material, s.thickness)
        import json
        try:
            stored_prices = json.loads(profile.custom_prices or '{}')
        except Exception:
            stored_prices = {}
        current_price = stored_prices.get(key, str(ref_price))
        groups[category].append({
            'spec': s,
            'key': key,
            'ref_price': ref_price,
            'current_price': current_price,
        })

    # 参考档位腐蚀价汇总（用于页面顶部提示）
    tier = profile.pricing_tier
    tier_ref = TIER_PRICES.get(tier, TIER_PRICES[3])

    return render(request, 'merchant/member_pricing.html', {
        'profile': profile,
        'groups': groups,
        'tier': tier,
        'tier_ref': tier_ref,
    })


@login_required
@merchant_required
@staff_permission('member_manage')
def member_approve(request, profile_id):
    """审核通过会员（保留入口，供直接调用）"""
    merchant = get_merchant(request)
    profile = get_object_or_404(CustomerProfile, pk=profile_id, merchant=merchant)
    if request.method == 'POST':
        profile.registration_status = 'approved'
        profile.credit_limit = Decimal(request.POST.get('credit_limit', '10000'))
        profile.save(update_fields=['registration_status', 'credit_limit'])
        profile.user.is_approved = True
        profile.user.save(update_fields=['is_approved'])
        messages.success(request, f'已通过 {profile.real_name or profile.user.phone} 的注册申请')
    return redirect('member_list')


@login_required
@merchant_required
@staff_permission('member_manage')
def member_reject(request, profile_id):
    """拒绝会员注册"""
    merchant = get_merchant(request)
    profile = get_object_or_404(CustomerProfile, pk=profile_id, merchant=merchant)
    if request.method == 'POST':
        profile.registration_status = 'rejected'
        profile.rejection_reason = request.POST.get('rejection_reason', '')
        profile.save(update_fields=['registration_status', 'rejection_reason'])
        messages.info(request, '已拒绝该注册申请')
    return redirect('member_list')


@login_required
@merchant_required
@staff_permission('member_manage')
def member_adjust_credit(request, profile_id):
    """调整会员信用额度"""
    merchant = get_merchant(request)
    profile = get_object_or_404(CustomerProfile, pk=profile_id, merchant=merchant, registration_status='approved')
    if request.method == 'POST':
        new_limit = Decimal(request.POST.get('credit_limit', '0'))
        profile.credit_limit = new_limit
        # 确保已用额度不超过新额度
        if profile.credit_used > new_limit:
            profile.credit_used = new_limit
        profile.save(update_fields=['credit_limit', 'credit_used'])
        messages.success(request, '信用额度已调整')
    return redirect('member_list')


# ==================== 工厂管理 ====================

@login_required
@merchant_required
def factory_list(request):
    """工厂列表"""
    merchant = get_merchant(request)
    factories = merchant.factories.all()
    return render(request, 'merchant/factory_list.html', {'factories': factories})


@login_required
@merchant_required
def factory_add(request):
    """添加工厂"""
    merchant = get_merchant(request)
    if request.method == 'POST':
        Factory.objects.create(
            merchant=merchant,
            name=request.POST.get('name'),
            address=request.POST.get('address'),
            contact_person=request.POST.get('contact_person'),
            contact_phone=request.POST.get('contact_phone'),
        )
        messages.success(request, '工厂已添加')
        return redirect('factory_list')
    return render(request, 'merchant/factory_form.html', {'title': '添加工厂'})


@login_required
@merchant_required
def factory_edit(request, pk):
    """编辑工厂"""
    merchant = get_merchant(request)
    factory = get_object_or_404(Factory, pk=pk, merchant=merchant)
    if request.method == 'POST':
        factory.name = request.POST.get('name', factory.name)
        factory.address = request.POST.get('address', factory.address)
        factory.contact_person = request.POST.get('contact_person', factory.contact_person)
        factory.contact_phone = request.POST.get('contact_phone', factory.contact_phone)
        factory.is_active = request.POST.get('is_active') == 'on'
        factory.save()
        messages.success(request, '工厂信息已更新')
        return redirect('factory_list')
    return render(request, 'merchant/factory_form.html', {'factory': factory, 'title': '编辑工厂'})


@login_required
@merchant_required
def factory_detail(request, pk):
    """工厂详情页：设备状态 + 库存管理"""
    merchant = get_merchant(request)
    factory = get_object_or_404(Factory, pk=pk, merchant=merchant)

    # 确保设备状态记录存在
    equipment_status, _ = FactoryEquipmentStatus.objects.get_or_create(
        factory=factory,
        defaults={'status_text': '一切正常'}
    )

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_equipment':
            equipment_status.status_text = request.POST.get('status_text', '一切正常')
            equipment_status.notes = request.POST.get('notes', '')
            equipment_status.save()
            messages.success(request, '设备状态已更新')

        elif action == 'add_inventory':
            FactoryInventory.objects.create(
                factory=factory,
                name=request.POST.get('name', ''),
                category=request.POST.get('category', ''),
                quantity=request.POST.get('quantity', ''),
                unit=request.POST.get('unit', ''),
                notes=request.POST.get('notes', '')
            )
            messages.success(request, '库存项已添加')

        elif action == 'update_inventory':
            inv_id = request.POST.get('inventory_id')
            inv = get_object_or_404(FactoryInventory, pk=inv_id, factory=factory)
            inv.name = request.POST.get('name', inv.name)
            inv.category = request.POST.get('category', inv.category)
            inv.quantity = request.POST.get('quantity', inv.quantity)
            inv.unit = request.POST.get('unit', inv.unit)
            inv.notes = request.POST.get('notes', inv.notes)
            inv.save()
            messages.success(request, '库存项已更新')

        elif action == 'delete_inventory':
            inv_id = request.POST.get('inventory_id')
            inv = get_object_or_404(FactoryInventory, pk=inv_id, factory=factory)
            inv.delete()
            messages.success(request, '库存项已删除')

        return redirect('factory_detail', pk=pk)

    inventories = factory.inventories.all()
    return render(request, 'merchant/factory_detail.html', {
        'factory': factory,
        'equipment_status': equipment_status,
        'inventories': inventories,
    })


@login_required
@merchant_required
def export_inventory(request, pk):
    """导出工厂库存盘点表Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    merchant = get_merchant(request)
    factory = get_object_or_404(Factory, pk=pk, merchant=merchant)
    inventories = factory.inventories.all().order_by('category', 'name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '库存盘点表'

    # ===== 样式定义 =====
    title_font = Font(name='微软雅黑', size=16, bold=True)
    subtitle_font = Font(name='微软雅黑', size=10, color='666666')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    normal_font = Font(name='微软雅黑', size=10)
    category_font = Font(name='微软雅黑', size=11, bold=True, color='1a56db')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    category_fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')

    # ===== 第1行：标题 =====
    ws.merge_cells('A1:H1')
    ws['A1'] = f'{factory.name} — 库存盘点表'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 32

    # ===== 第2行：信息 =====
    ws.merge_cells('A2:H2')
    ws['A2'] = f'导出时间：{timezone.now().strftime("%Y年%m月%d日 %H:%M")}    共 {inventories.count()} 项'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # 空行
    ws.row_dimensions[3].height = 8

    # ===== 第4行：表头 =====
    headers = ['序号', '名称', '分类', '系统数量', '单位', '盘点数量', '差异', '备注']
    col_widths = [6, 38, 14, 14, 8, 14, 14, 30]
    for idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[4].height = 26

    # ===== 数据行 =====
    current_row = 5
    current_category = None
    seq = 0

    for inv in inventories:
        # 分类分隔行（新分类时插入）
        if inv.category != current_category:
            if current_category is not None:
                current_row += 1  # 空一行
            ws.merge_cells(f'A{current_row}:H{current_row}')
            cell = ws.cell(row=current_row, column=1, value=f'▸ {inv.category or "未分类"}')
            cell.font = category_font
            cell.fill = category_fill
            cell.alignment = left_align
            cell.border = thin_border
            for c in range(2, 9):
                ws.cell(row=current_row, column=c).border = thin_border
            ws.row_dimensions[current_row].height = 24
            current_row += 1
            current_category = inv.category

        seq += 1
        row_data = [
            seq,
            inv.name,
            inv.category or '',
            inv.quantity or '',
            inv.unit or '',
            '',  # 盘点数量（空白供手写）
            '',  # 差异（空白）
            inv.notes or '',
        ]

        for idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if idx == 1:
                cell.alignment = center_align
            elif idx in (6, 7):
                cell.fill = PatternFill(start_color='fffbeb', end_color='fffbeb', fill_type='solid')
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # ===== 底部签名区 =====
    current_row += 1
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.cell(row=current_row, column=1, value='盘点人签字：__________________')
    ws.cell(row=current_row, column=1).font = normal_font
    ws.cell(row=current_row, column=1).alignment = left_align

    ws.merge_cells(f'E{current_row}:H{current_row}')
    ws.cell(row=current_row, column=5, value='审核人签字：__________________')
    ws.cell(row=current_row, column=5).font = normal_font
    ws.cell(row=current_row, column=5).alignment = left_align

    # ===== 冻结窗格 & 打印设置 =====
    ws.freeze_panes = 'A5'
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.print_options.gridLines = False

    # ===== 输出 =====
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"库存盘点表_{factory.name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ==================== 商品规格管理 ====================

@login_required
@merchant_required
def spec_list(request):
    """商品规格列表"""
    merchant = get_merchant(request)
    # 平台预设规格
    preset_specs = ProductSpec.objects.filter(is_platform_preset=True)
    # 商家自己的规格状态（上下架）
    merchant_specs = {s.product_name+'_'+s.material+'_'+s.thickness: s for s in ProductSpec.objects.filter(merchant=merchant)}
    specs = []
    from utils.pricing_tiers import is_etching_product, TIER_PRICES
    for ps in preset_specs:
        key = ps.product_name+'_'+ps.material+'_'+ps.thickness
        ms = merchant_specs.get(key)
        # 价格显示文本
        if is_etching_product(ps.product_name):
            price_1 = TIER_PRICES[1].get(ps.thickness, '-')
            price_2 = TIER_PRICES[2].get(ps.thickness, '-')
            price_3 = TIER_PRICES[3].get(ps.thickness, '-')
            price_display = f'1档¥{price_1} / 2档¥{price_2} / 3档¥{price_3}'
        else:
            price_display = f'¥{ps.unit_price}'
        specs.append({
            'preset': ps,
            'merchant': ms,
            'is_active': ms.is_active if ms else True,
            'price_display': price_display,
        })
    custom_requests = CustomSpecRequest.objects.filter(merchant=merchant).order_by('-created_at')[:5]
    return render(request, 'merchant/spec_list.html', {
        'specs': specs,
        'custom_requests': custom_requests,
    })


@login_required
@merchant_required
def spec_toggle(request):
    """上架/下架规格"""
    if request.method == 'POST':
        merchant = get_merchant(request)
        spec_id = request.POST.get('spec_id')
        preset = get_object_or_404(ProductSpec, pk=spec_id, is_platform_preset=True)
        # 查找或创建商家规格记录
        spec, created = ProductSpec.objects.get_or_create(
            merchant=merchant, product_name=preset.product_name, material=preset.material,
            thickness=preset.thickness,
            defaults={'unit_price': preset.unit_price, 'is_platform_preset': False, 'is_active': True}
        )
        spec.is_active = not spec.is_active
        spec.save(update_fields=['is_active'])
        return JsonResponse({'success': True, 'is_active': spec.is_active})
    return JsonResponse({'success': False}, status=400)


@login_required
@merchant_required
def spec_custom_request(request):
    """申请非标规格"""
    merchant = get_merchant(request)
    if request.method == 'POST':
        CustomSpecRequest.objects.create(
            merchant=merchant,
            material=request.POST.get('material'),
            process_type=request.POST.get('process_type', ''),
            thickness=request.POST.get('thickness'),
            description=request.POST.get('description', ''),
        )
        messages.success(request, '非标规格申请已提交，等待平台审核')
        return redirect('spec_list')
    return render(request, 'merchant/spec_custom_request.html')


# ==================== 订单管理 ====================

@login_required
@merchant_required
def merchant_orders(request):
    """商家订单列表"""
    merchant = get_merchant(request)
    status_filter = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    customer_query = request.GET.get('customer')
    statement_status = request.GET.get('statement_status', '')
    orders = merchant.orders.filter(is_submitted=True).select_related('customer__customer_profile').prefetch_related('items', 'plate_layout', 'statement', 'complaints')
    if status_filter:
        orders = orders.filter(status=status_filter)
    if date_from:
        orders = orders.filter(created_at__date__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__date__lte=date_to)
    if customer_query:
        orders = orders.filter(customer__phone__icontains=customer_query)
    # 对账状态筛选
    if statement_status == 'unsettled':
        orders = orders.filter(status='received', is_settled=False, statement__isnull=True)
    elif statement_status == 'statemented':
        orders = orders.filter(statement__isnull=False, is_settled=False)
    elif statement_status == 'settled':
        orders = orders.filter(is_settled=True)
    orders = orders.order_by('-created_at')
    
    # 【新增】计算每个订单的时效状态 + 物流摘要
    from django.utils import timezone
    from utils.kuaidi100 import query_tracking, format_tracking_data
    now = timezone.now()
    for order in orders:
        order.sla_status = None
        order.sla_label = '-'
        order.sla_class = ''
        
        # 物流摘要：已发货/已收货的订单预加载
        order.tracking_summary = None
        if order.status in ('shipped', 'received') and order.tracking_number:
            # 优先使用缓存
            if order.tracking_status and order.tracking_last_context:
                order.tracking_summary = order.get_tracking_summary()
            else:
                # 缓存为空，查询API并更新缓存
                try:
                    result = query_tracking(order.tracking_number, order.tracking_company or None)
                    if result['success']:
                        tracking_data = format_tracking_data(result['data'])
                        order.update_tracking_cache(tracking_data)
                        order.tracking_summary = order.get_tracking_summary()
                except Exception:
                    pass
        
        # 客服处理时效：pending_confirm 状态
        if order.status == 'pending_confirm' and order.file_uploaded_at:
            elapsed = (now - order.file_uploaded_at).total_seconds() / 60
            if order.customer_service_processed_at:
                order.sla_status = 'done'
                order.sla_label = f'已处理'
                order.sla_class = 'bg-secondary'
            elif elapsed > 30:
                order.sla_status = 'overdue'
                order.sla_label = f'已超{int(elapsed - 30)}分'
                order.sla_class = 'bg-danger'
            elif elapsed > 15:
                order.sla_status = 'warning'
                order.sla_label = f'剩{int(30 - elapsed)}分'
                order.sla_class = 'bg-warning text-dark'
            else:
                order.sla_status = 'normal'
                order.sla_label = '正常'
                order.sla_class = 'bg-success'
        
        # 工厂下载时效：in_production 状态
        elif order.status == 'in_production' and order.factory_notified_at:
            if order.factory_downloaded_at:
                order.sla_status = 'done'
                order.sla_label = '已下载'
                order.sla_class = 'bg-secondary'
            else:
                elapsed = (now - order.factory_notified_at).total_seconds() / 60
                if elapsed > 30:
                    order.sla_status = 'overdue'
                    order.sla_label = f'已超{int(elapsed - 30)}分'
                    order.sla_class = 'bg-danger'
                elif elapsed > 15:
                    order.sla_status = 'warning'
                    order.sla_label = f'剩{int(30 - elapsed)}分'
                    order.sla_class = 'bg-warning text-dark'
                else:
                    order.sla_status = 'normal'
                    order.sla_label = '正常'
                    order.sla_class = 'bg-success'
    
    return render(request, 'merchant/orders.html', {
        'orders': orders,
        'status_choices': Order.STATUS_CHOICES,
        'status_filter': status_filter,
        'statement_status': statement_status,
    })


@login_required
@merchant_required
def merchant_order_detail(request, order_id):
    """商家订单详情"""
    merchant = get_merchant(request)
    order = get_object_or_404(Order, pk=order_id, merchant=merchant)
    factories = merchant.factories.filter(is_active=True)
    designers = StaffProfile.objects.filter(merchant=merchant, is_active=True, role__name='designer')
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'communicate':
            content = request.POST.get('content', '')
            if content:
                CommunicationLog.objects.create(order=order, sender=request.user, content=content)
                messages.success(request, '沟通记录已添加')
        elif action == 'audit':
            # 审核通过 -> 待设计/已确认
            from django.utils import timezone
            order.transition_status('design_confirmed', operator=request.user, remark='商家审核通过')
            order.customer_service_processed_at = timezone.now()  # 【新增】记录客服处理时间
            # 自动拼版：为该订单生成拼版建议
            auto_generate_plate_layout_for_order(order)
            order.plate_status = 'auto_generated'
            order.save(update_fields=['plate_status', 'customer_service_processed_at'])
            messages.success(request, '订单已审核通过，系统已自动拼版，请设计师确认')
        elif action == 'assign_design':
            designer_id = request.POST.get('designer_id')
            if designer_id:
                order.design_assigned_to_id = designer_id
                order.save(update_fields=['design_assigned_to'])
                order.transition_status('design_confirmed', operator=request.user, remark='已分配设计岗')
                messages.success(request, '已分配给设计人员')
        elif action == 'arrange_production':
            factory_id = request.POST.get('factory_id')
            cycle = request.POST.get('production_cycle')
            if factory_id and cycle:
                # 检查是否已完成拼版（兼容旧版 PlateLayout 和 新版 PlateBatch）
                has_layout = False
                plate_layout = getattr(order, 'plate_layout', None)
                if plate_layout and plate_layout.layout_data:
                    has_layout = True
                if order.plate_batch_items.exists():
                    has_layout = True
                if not has_layout:
                    messages.error(request, '该订单尚未完成拼版，请先安排设计师拼版')
                    return redirect('merchant_order_detail', order_id=order_id)
                from django.utils import timezone
                order.factory_id = factory_id
                order.production_cycle = int(cycle)
                order.factory_notified_at = timezone.now()  # 【新增】记录工厂通知时间
                order.save(update_fields=['factory', 'production_cycle', 'factory_notified_at'])
                order.transition_status('in_production', operator=request.user, remark='已安排生产')
                # 同步更新 PlateBatch
                for pbi in order.plate_batch_items.all():
                    pb = pbi.plate_batch
                    if pb.status in ('auto_generated', 'confirmed'):
                        pb.status = 'in_production'
                        pb.factory_id = factory_id
                        pb.save(update_fields=['status', 'factory'])
                messages.success(request, '生产已安排，工厂将在下个整点收到订单')
        elif action == 'reject':
            reason = request.POST.get('rejection_reason', '')
            order.rejection_reason = reason
            order.save(update_fields=['rejection_reason'])
            order.transition_status('info_error', operator=request.user, remark=f'无法生产: {reason}')
            messages.info(request, '订单已驳回')
        elif action == 'ship':
            from django.utils import timezone
            order.tracking_number = request.POST.get('tracking_number', '')
            order.tracking_company = request.POST.get('tracking_company', '')
            order.shipped_at = timezone.now()
            order.save(update_fields=['tracking_number', 'tracking_company', 'shipped_at'])
            order.transition_status('shipped', operator=request.user, remark='已发货')
            messages.success(request, '订单已发货')
            # 【新增】尝试订阅快递100推送
            if order.tracking_number and order.tracking_company:
                from utils.kuaidi100 import get_company_code, subscribe_tracking
                company_code = get_company_code(order.tracking_company)
                if company_code:
                    subscribe_result = subscribe_tracking(
                        order.tracking_number,
                        company_code,
                        phone_tail=order.delivery_address.phone[-4:] if order.delivery_address and order.delivery_address.phone else None
                    )
                    if subscribe_result['success']:
                        logger.info("快递100订阅成功: order=%s tracking=%s", order.id, order.tracking_number)
                    else:
                        logger.warning("快递100订阅失败: order=%s tracking=%s msg=%s", order.id, order.tracking_number, subscribe_result['message'])
                else:
                    logger.warning("未找到快递公司编码: %s", order.tracking_company)
        elif action == 'mark_paid':
            old_status = order.status
            order.transition_status('paid', operator=request.user, remark='线下支付已确认')
            # 只有从"待支付"转为"已支付"时才扣除额度，避免重复扣
            if old_status == 'pending_payment':
                profile = order.customer.customer_profile
                if profile.credit_remaining >= order.total_amount:
                    profile.credit_used += order.total_amount
                    profile.save(update_fields=['credit_used'])
                else:
                    messages.warning(request, '客户信用额度不足，订单已标记为已支付但额度未扣减')
            messages.success(request, '已标记为已支付')
        elif action == 'confirm_payment':
            # 确认收款并释放额度（订单层面直接处理）
            if order.status not in ('paid', 'received', 'shipped', 'in_production'):
                messages.error(request, '该订单当前状态不支持确认收款')
            elif order.is_settled:
                messages.info(request, '该订单已结清，无需重复操作')
            elif order.statement and order.statement.status != 'settled':
                messages.warning(
                    request,
                    f'该订单已关联对账单 {order.statement.sn}，请通过对账单页面进行结清操作'
                )
            else:
                with transaction.atomic():
                    order.is_settled = True
                    order.save(update_fields=['is_settled'])
                    profile = order.customer.customer_profile
                    profile.credit_used -= order.total_amount
                    if profile.credit_used < 0:
                        profile.credit_used = 0
                    profile.save(update_fields=['credit_used'])
                messages.success(
                    request,
                    f'已确认收款并释放额度 ¥{order.total_amount}，'
                    f'客户 {order.customer.phone} 当前已用额度 ¥{profile.credit_used} / ¥{profile.credit_limit}'
                )
        elif action == 'reupload_file':
            # 设计师重新上传处理后的文件
            item_id = request.POST.get('item_id')
            item = get_object_or_404(OrderItem, pk=item_id, order=order)
            if request.FILES.get('pdf_file'):
                file = request.FILES['pdf_file']
                ext = os.path.splitext(file.name)[1].lower()
                if ext not in ('.pdf', '.ai'):
                    messages.error(request, '仅支持PDF或AI格式文件')
                    return redirect('merchant_order_detail', order_id=order_id)
                # 保存新文件
                import uuid
                filename = f"order_files/{order.customer.id}/{uuid.uuid4().hex}{ext}"
                path = default_storage.save(filename, file)
                item.file = path
                item.save(update_fields=['file'])
                # 自动重新拼版
                try:
                    auto_generate_plate_layout_for_order(order)
                    order.plate_status = 'auto_generated'
                    order.save(update_fields=['plate_status'])
                    messages.success(request, f'文件已重新上传，系统已自动重新拼版，请设计师确认')
                except Exception as e:
                    messages.success(request, f'文件已重新上传，但自动拼版失败，请手动处理')
            else:
                messages.error(request, '请选择PDF文件')
        elif action == 'update_red_box':
            messages.error(request, '商户无权修改客户红框尺寸')
            return redirect('merchant_order_detail', order_id=order_id)
        return redirect('merchant_order_detail', order_id=order_id)

    # GET: 生成预检报告供商户/设计师参考
    import os
    from django.conf import settings
    from utils.pdf_preflight import preflight_pdf, generate_preflight_report_html
    preflight_reports = []
    has_file_issues = False
    for item in order.items.all():
        if item.file:
            try:
                file_path = os.path.join(settings.MEDIA_ROOT, item.file.name)
                if os.path.exists(file_path):
                    report = preflight_pdf(file_path, min_line_width_mm=0.12)
                    preflight_reports.append({
                        'item_id': str(item.id),
                        'item_label': f"{item.get_product_name_display()} #{str(item.id)[:8]}",
                        'report_html': generate_preflight_report_html(report),
                        'pass': report['pass'],
                    })
                    if not report['pass']:
                        has_file_issues = True
            except Exception:
                pass

    # 【新增】为没有预览图的旧订单生成预览图
    for item in order.items.all():
        if item.file and not item.preview_image:
            try:
                from utils.pdf_processor import generate_pdf_preview
                preview_filename = f"previews/{item.id}.png"
                preview_path = os.path.join(settings.MEDIA_ROOT, preview_filename)
                if not os.path.exists(preview_path):
                    generate_pdf_preview(item.file.name, preview_filename, dpi=150)
                if os.path.exists(preview_path):
                    item.preview_image = preview_filename
                    item.save(update_fields=['preview_image'])
            except Exception:
                pass

    # 查询快递100物流轨迹
    tracking_data = None
    if order.tracking_number:
        from utils.kuaidi100 import query_tracking, format_tracking_data
        result = query_tracking(order.tracking_number)
        if result['success']:
            tracking_data = format_tracking_data(result['data'])

    # 计算时效追踪数据（分钟）
    from django.utils import timezone
    now = timezone.now()
    cs_elapsed = None
    cs_overdue = None
    cs_remaining = None
    if order.file_uploaded_at and not order.customer_service_processed_at:
        cs_elapsed = int((now - order.file_uploaded_at).total_seconds() / 60)
        if cs_elapsed > 30:
            cs_overdue = cs_elapsed - 30
        elif cs_elapsed > 15:
            cs_remaining = 30 - cs_elapsed
    fd_elapsed = None
    fd_overdue = None
    fd_remaining = None
    if order.factory_notified_at and not order.factory_downloaded_at:
        fd_elapsed = int((now - order.factory_notified_at).total_seconds() / 60)
        if fd_elapsed > 30:
            fd_overdue = fd_elapsed - 30
        elif fd_elapsed > 15:
            fd_remaining = 30 - fd_elapsed

    return render(request, 'merchant/order_detail.html', {
        'order': order,
        'factories': factories,
        'designers': designers,
        'preflight_reports': preflight_reports,
        'has_file_issues': has_file_issues,
        'tracking_data': tracking_data,
        'cs_elapsed': cs_elapsed,
        'cs_overdue': cs_overdue,
        'cs_remaining': cs_remaining,
        'fd_elapsed': fd_elapsed,
        'fd_overdue': fd_overdue,
        'fd_remaining': fd_remaining,
    })


@login_required
@merchant_required
def upload_production_photo(request, order_id):
    """上传生产/质检照片"""
    merchant = get_merchant(request)
    order = get_object_or_404(Order, pk=order_id, merchant=merchant)
    if request.method == 'POST' and request.FILES.get('photo'):
        ProductionPhoto.objects.create(
            order=order,
            photo_type=request.POST.get('photo_type', 'front'),
            image=request.FILES['photo'],
            uploaded_by=request.user,
        )
        messages.success(request, '照片已上传')
    return redirect('merchant_order_detail', order_id=order_id)


@login_required
@merchant_required
@staff_permission('order_audit')
def batch_process_orders(request):
    """批量处理订单（二级页面）：支持审核、拼版确认、安排生产、上传文件"""
    merchant = get_merchant(request)
    factories = merchant.factories.filter(is_active=True)

    if request.method == 'POST':
        order_ids_str = request.POST.get('order_ids', '')
        order_ids = [oid.strip() for oid in order_ids_str.split(',') if oid.strip()]
        if not order_ids:
            messages.warning(request, '未选择任何订单')
            return redirect('merchant_orders')

        success_count = 0
        error_msgs = []

        for oid in order_ids:
            try:
                order = Order.objects.get(pk=oid, merchant=merchant, is_submitted=True)
            except Order.DoesNotExist:
                continue

            action = None
            if request.POST.get(f'do_audit_{oid}'):
                action = 'audit'
            elif request.POST.get(f'do_confirm_{oid}'):
                action = 'confirm_plate'
            elif request.POST.get(f'do_arrange_{oid}'):
                action = 'arrange_production'
            elif request.POST.get(f'do_ship_{oid}'):
                action = 'ship'

            # 文件上传单独检测（不需要 checkbox）
            file_uploaded = False
            for item in order.items.all():
                file_key = f'file_{oid}_{item.id}'
                if file_key in request.FILES:
                    uploaded = request.FILES[file_key]
                    ext = os.path.splitext(uploaded.name)[1].lower()
                    if ext not in ('.pdf', '.ai'):
                        error_msgs.append(f'{order.sn}: 仅支持PDF或AI格式')
                        continue
                    import uuid
                    filename = f"order_files/{order.customer.id}/{uuid.uuid4().hex}{ext}"
                    path = default_storage.save(filename, uploaded)
                    item.file = path
                    item.save(update_fields=['file'])
                    file_uploaded = True
            if file_uploaded:
                try:
                    auto_generate_plate_layout_for_order(order)
                    order.plate_status = 'auto_generated'
                    order.save(update_fields=['plate_status'])
                except Exception:
                    pass
                success_count += 1
                # 如果同时有其他 action，也继续执行

            if not action:
                continue

            try:
                if action == 'audit':
                    if order.status == 'pending_confirm':
                        from django.utils import timezone
                        order.transition_status('design_confirmed', operator=request.user, remark='批量审核通过')
                        order.customer_service_processed_at = timezone.now()
                        auto_generate_plate_layout_for_order(order)
                        order.plate_status = 'auto_generated'
                        order.save(update_fields=['plate_status', 'customer_service_processed_at'])
                        success_count += 1
                    else:
                        error_msgs.append(f'{order.sn}: 当前状态不支持审核通过')

                elif action == 'confirm_plate':
                    if order.plate_status == 'auto_generated':
                        has_layout = False
                        layout = getattr(order, 'plate_layout', None)
                        if layout and layout.layout_data:
                            has_layout = True
                        if order.plate_batch_items.exists():
                            has_layout = True
                        if has_layout:
                            order.plate_status = 'confirmed'
                            default_factory = factories.first()
                            if default_factory:
                                order.factory = default_factory
                            order.save(update_fields=['plate_status', 'factory'])
                            order.transition_status('in_production', operator=request.user, remark='批量确认拼版，自动下发工厂生产')
                            # 同步更新关联的 PlateBatch
                            for pbi in order.plate_batch_items.all():
                                pb = pbi.plate_batch
                                if pb.status == 'auto_generated':
                                    pb.status = 'confirmed'
                                    pb.factory = default_factory
                                    pb.save(update_fields=['status', 'factory'])
                            success_count += 1
                        else:
                            error_msgs.append(f'{order.sn}: 尚未生成拼版数据')
                    else:
                        error_msgs.append(f'{order.sn}: 拼版状态不支持确认')

                elif action == 'arrange_production':
                    if order.status in ('paid', 'design_confirmed'):
                        factory_id = request.POST.get(f'factory_{oid}')
                        cycle = request.POST.get(f'cycle_{oid}')
                        if factory_id and cycle:
                            has_layout = False
                            plate_layout = getattr(order, 'plate_layout', None)
                            if plate_layout and plate_layout.layout_data:
                                has_layout = True
                            if order.plate_batch_items.exists():
                                has_layout = True
                            if not has_layout:
                                error_msgs.append(f'{order.sn}: 该订单尚未完成拼版')
                                continue
                            from django.utils import timezone
                            order.factory_id = factory_id
                            order.production_cycle = int(cycle)
                            order.factory_notified_at = timezone.now()
                            order.save(update_fields=['factory', 'production_cycle', 'factory_notified_at'])
                            order.transition_status('in_production', operator=request.user, remark='批量安排生产')
                            # 同步更新 PlateBatch
                            for pbi in order.plate_batch_items.all():
                                pb = pbi.plate_batch
                                if pb.status in ('auto_generated', 'confirmed'):
                                    pb.status = 'in_production'
                                    pb.factory_id = factory_id
                                    pb.save(update_fields=['status', 'factory'])
                            success_count += 1
                        else:
                            error_msgs.append(f'{order.sn}: 请选择工厂和发货时效')
                    else:
                        error_msgs.append(f'{order.sn}: 当前状态不支持安排生产')

                elif action == 'ship':
                    if order.status == 'in_production':
                        tracking = request.POST.get(f'tracking_{oid}', '')
                        company = request.POST.get(f'company_{oid}', '')
                        order.tracking_number = tracking
                        order.tracking_company = company
                        order.shipped_at = timezone.now()
                        order.save(update_fields=['tracking_number', 'tracking_company', 'shipped_at'])
                        order.transition_status('shipped', operator=request.user, remark='批量发货')
                        success_count += 1
                    else:
                        error_msgs.append(f'{order.sn}: 当前状态不支持发货')

            except Exception as e:
                error_msgs.append(f'{order.sn}: 处理失败 ({str(e)})')

        if success_count > 0:
            messages.success(request, f'成功处理 {success_count} 个订单')
        if error_msgs:
            for msg in error_msgs[:5]:
                messages.warning(request, msg)
            if len(error_msgs) > 5:
                messages.warning(request, f'还有 {len(error_msgs) - 5} 个订单处理失败')
        return redirect('merchant_orders')

    # GET
    order_ids_str = request.GET.get('order_ids', '')
    order_ids = [oid.strip() for oid in order_ids_str.split(',') if oid.strip()]
    if not order_ids:
        messages.warning(request, '未选择任何订单')
        return redirect('merchant_orders')

    orders = Order.objects.filter(
        pk__in=order_ids, merchant=merchant, is_submitted=True
    ).prefetch_related('items', 'plate_layout')

    # 按状态分组统计
    status_counts = {}
    for o in orders:
        status_counts[o.status] = status_counts.get(o.status, 0) + 1

    return render(request, 'merchant/batch_process_orders.html', {
        'orders': orders,
        'factories': factories,
        'status_counts': status_counts,
        'order_ids_str': order_ids_str,
    })


# ==================== 拼版工具 ====================

@login_required
@merchant_required
def plate_layout_orders(request):
    """设计岗：获取待拼版/待确认订单"""
    merchant = get_merchant(request)
    # 需要拼版的订单状态：design_confirmed（商家审核后）或 paid（客户信用额度支付后直接到账）
    eligible_statuses = ['design_confirmed', 'paid']
    # 设计岗只能看到自己的，管理员看到所有
    if request.user.user_type == 'merchant_staff' and request.user.staff_profile.role.name == 'designer':
        orders = merchant.orders.filter(design_assigned_to=request.user, status__in=eligible_statuses)
    else:
        orders = merchant.orders.filter(status__in=eligible_statuses)
    orders = orders.order_by('-created_at')
    # 分组：待确认拼版（系统自动生成） vs 待手动拼版（还未生成）
    pending_confirm = orders.filter(plate_status='auto_generated')
    pending_work = orders.filter(plate_status='none')
    return render(request, 'merchant/plate_layout_orders.html', {
        'pending_confirm': pending_confirm,
        'pending_work': pending_work,
    })


@login_required
@merchant_required
def plate_layout_work(request, order_id):
    """拼版工作台（支持多种拼版算法切换）"""
    merchant = get_merchant(request)
    order = get_object_or_404(Order, pk=order_id, merchant=merchant)
    items = order.items.all()
    
    # 算法选择
    ALGORITHMS = [
        ('shelf', '原生日志算法 (Shelf)'),
        ('maxrects', 'MaxRects（推荐·利用率最高）'),
        ('guillotine', 'Guillotine（平衡型）'),
        ('skyline', 'Skyline（快速型）'),
    ]
    algorithm = request.GET.get('algorithm', 'maxrects')
    if algorithm not in [a[0] for a in ALGORITHMS]:
        algorithm = 'maxrects'
    
    # 获取尺寸信息用于拼版算法
    rects = []
    for item in items:
        if item.length_mm and item.width_mm:
            rects.append({
                'id': str(item.id),
                'width': float(item.length_mm),
                'height': float(item.width_mm),
                'label': f"{item.get_material_display()}",
            })
    # 拼版建议
    suggestion = None
    if rects:
        suggestion = calculate_plate_layout_rectpack(rects, plate_width=600, plate_height=1000, algorithm=algorithm)
    layout = getattr(order, 'plate_layout', None)

    # 解析已保存的拼版数据（JSON字符串 -> dict）用于模板展示
    layout_data_parsed = {}
    if layout and layout.layout_data:
        try:
            layout_data_parsed = json.loads(layout.layout_data)
        except Exception:
            layout_data_parsed = {}
    # 如果已有保存的拼版数据，优先用它作为画布展示
    if layout_data_parsed:
        suggestion = layout_data_parsed

    # 计算画布缩放比例（画布固定高度500px，宽度按col-md-8约750px估算）
    CANVAS_WIDTH = 750
    CANVAS_HEIGHT = 500
    scaled_rectangles = []
    scaled_plate_width = 0
    scaled_plate_height = 0
    scale = 1.0
    
    # 预生成每个订单项的PDF预览图（用于拼版画布显示实际内容，带版类视觉效果）
    item_previews = {}
    for item in items:
        if item.file:
            try:
                file_path = os.path.join(settings.MEDIA_ROOT, item.file.name)
                if os.path.exists(file_path):
                    preview_filename = f"previews/plate_{order.id}_{item.id}.png"
                    from utils.plate_preview_effects import generate_effect_preview
                    preview_path = generate_effect_preview(
                        file_path, preview_filename,
                        product_name=item.product_name,
                        plate_type_key=getattr(item, 'plate_type', None),
                        dpi=150
                    )
                    if preview_path:
                        item_previews[str(item.id)] = settings.MEDIA_URL + preview_filename
            except Exception:
                pass
    
    if suggestion and suggestion.get('plate_width') and suggestion.get('plate_height'):
        pw = float(suggestion['plate_width'])
        ph = float(suggestion['plate_height'])
        scale_w = CANVAS_WIDTH / pw
        scale_h = CANVAS_HEIGHT / ph
        scale = min(scale_w, scale_h)
        scaled_plate_width = round(pw * scale, 1)
        scaled_plate_height = round(ph * scale, 1)
        for rect in suggestion.get('rectangles', []):
            rect_id = rect.get('id', '')
            scaled_rectangles.append({
                'id': rect_id,
                'x': round(float(rect.get('x', 0)) * scale, 1),
                'y': round(float(rect.get('y', 0)) * scale, 1),
                'width': round(float(rect.get('width', 0)) * scale, 1),
                'height': round(float(rect.get('height', 0)) * scale, 1),
                'label': rect.get('label', ''),
                'preview_url': item_previews.get(rect_id, ''),
            })

    if request.method == 'POST':
        action = request.POST.get('action', 'save')

        if action == 'save':
            data = json.loads(request.POST.get('layout_data', '{}'))
            note = request.POST.get('designer_note', '')
            layout_json = json.dumps(data, ensure_ascii=False)
            if layout:
                layout.layout_data = layout_json
                layout.designer_note = note
                layout.designer = request.user
                layout.save()
            else:
                PlateLayout.objects.create(order=order, layout_data=layout_json, designer_note=note, designer=request.user)
            messages.success(request, '拼版布局已保存')
            return redirect('plate_layout_orders')

        elif action == 'confirm_plate':
            # 设计师确认拼版：保存当前布局，确认拼版，自动下发到工厂
            data = json.loads(request.POST.get('layout_data', '{}'))
            note = request.POST.get('designer_note', '')
            layout_json = json.dumps(data, ensure_ascii=False)
            if layout:
                layout.layout_data = layout_json
                layout.designer_note = note
                layout.designer = request.user
                layout.save()
            else:
                PlateLayout.objects.create(order=order, layout_data=layout_json, designer_note=note, designer=request.user)
            # 确认拼版状态
            order.plate_status = 'confirmed'
            # 自动分配给第一个活跃工厂
            default_factory = merchant.factories.filter(is_active=True).first()
            if default_factory:
                order.factory = default_factory
            from django.utils import timezone
            order.factory_notified_at = timezone.now()
            order.save(update_fields=['plate_status', 'factory', 'factory_notified_at'])
            # 状态变为生产中，工厂立刻收到
            order.transition_status('in_production', operator=request.user, remark='设计师已确认拼版，自动下发工厂生产')
            messages.success(request, f'订单 {order.sn} 拼版已确认，已自动下发到工厂生产')
            return redirect('plate_layout_orders')

        elif action == 'reject_plate':
            reason = request.POST.get('reject_reason', '')
            order.plate_status = 'rejected'
            order.save(update_fields=['plate_status'])
            OrderStatusLog.objects.create(
                order=order, from_status=order.status, to_status=order.status,
                operator=request.user, remark=f'设计师驳回拼版: {reason}'
            )
            messages.warning(request, f'订单 {order.sn} 拼版已驳回，原因：{reason}')
            return redirect('plate_layout_orders')

    # PDF预检报告
    preflight_reports = []
    import os
    from django.conf import settings
    from utils.pdf_preflight import preflight_pdf, generate_preflight_report_html
    for item in items:
        if item.file:
            try:
                file_path = os.path.join(settings.MEDIA_ROOT, item.file.name)
                if os.path.exists(file_path):
                    report = preflight_pdf(file_path, min_line_width_mm=0.12)
                    preflight_reports.append({
                        'item_id': str(item.id),
                        'item_label': f"{item.get_product_name_display()} #{item.id[:8]}",
                        'report_html': generate_preflight_report_html(report),
                        'pass': report['pass'],
                    })
            except Exception as e:
                pass

    return render(request, 'merchant/plate_layout_work.html', {
        'order': order,
        'items': items,
        'suggestion': suggestion,
        'suggestion_json': json.dumps(suggestion, ensure_ascii=False) if suggestion else '{}',
        'layout': layout,
        'layout_data_parsed': layout_data_parsed,
        'preflight_reports': preflight_reports,
        'scaled_rectangles': scaled_rectangles,
        'scaled_plate_width': scaled_plate_width,
        'scaled_plate_height': scaled_plate_height,
        'algorithm': algorithm,
        'algorithms': ALGORITHMS,
    })


# ==================== 岗位权限与子账号 ====================

@login_required
@merchant_admin_required
def role_list(request):
    """岗位与权限管理"""
    merchant = get_merchant(request)
    roles = merchant.roles.all()
    return render(request, 'merchant/role_list.html', {'roles': roles})


@login_required
@merchant_admin_required
def role_edit(request, role_id):
    """编辑角色权限"""
    merchant = get_merchant(request)
    role = get_object_or_404(Role, pk=role_id, merchant=merchant)
    if request.method == 'POST':
        perms = request.POST.getlist('permissions')
        role.set_permissions(perms)
        role.save(update_fields=['permissions'])
        messages.success(request, '权限已更新')
        return redirect('role_list')
    all_permissions = [
        ('order_view', '查看订单'),
        ('order_audit', '审核订单'),
        ('order_production', '安排生产'),
        ('order_ship', '发货管理'),
        ('design_layout', '拼版设计'),
        ('member_manage', '会员管理'),
        ('factory_manage', '工厂管理'),
        ('spec_manage', '商品规格管理'),
        ('subaccount_manage', '子账号管理'),
        ('finance_manage', '财务管理'),
    ]
    current_perms = role.permissions or {}
    return render(request, 'merchant/role_edit.html', {
        'role': role,
        'all_permissions': all_permissions,
        'current_perms': current_perms,
    })


@login_required
@merchant_admin_required
def subaccount_list(request):
    """子账号管理"""
    merchant = get_merchant(request)
    staff = merchant.staff.filter(user__user_type='merchant_staff')
    return render(request, 'merchant/subaccount_list.html', {
        'staff': staff,
        'max_count': merchant.max_sub_accounts,
        'current_count': staff.count(),
    })


@login_required
@merchant_admin_required
def subaccount_add(request):
    """新增子账号"""
    merchant = get_merchant(request)
    if merchant.staff.filter(user__user_type='merchant_staff').count() >= merchant.max_sub_accounts:
        messages.error(request, f'子账号数量已达上限 ({merchant.max_sub_accounts})')
        return redirect('subaccount_list')
    roles = merchant.roles.all()
    if request.method == 'POST':
        phone = request.POST.get('phone')
        username = request.POST.get('username')
        password = request.POST.get('password')
        role_id = request.POST.get('role_id')
        if User.objects.filter(phone=phone).exists():
            messages.error(request, '该手机号已被注册')
        else:
            user = User.objects.create_user(
                username=username, phone=phone, password=password,
                user_type='merchant_staff', is_approved=True
            )
            StaffProfile.objects.create(
                user=user, merchant=merchant,
                role_id=role_id if role_id else None
            )
            messages.success(request, '子账号创建成功')
            return redirect('subaccount_list')
    return render(request, 'merchant/subaccount_form.html', {'roles': roles, 'title': '新增子账号'})


@login_required
@merchant_admin_required
def subaccount_edit(request, staff_id):
    """编辑子账号"""
    merchant = get_merchant(request)
    staff = get_object_or_404(StaffProfile, pk=staff_id, merchant=merchant)
    roles = merchant.roles.all()
    if request.method == 'POST':
        staff.is_active = request.POST.get('is_active') == 'on'
        staff.role_id = request.POST.get('role_id') or None
        staff.save()
        messages.success(request, '子账号已更新')
        return redirect('subaccount_list')
    return render(request, 'merchant/subaccount_form.html', {
        'staff': staff, 'roles': roles, 'title': '编辑子账号'
    })


# ==================== 工厂生产看板 ====================

@login_required
@merchant_required
def factory_production_board(request):
    """工厂生产看板 - 显示待生产/生产中/已完成的订单"""
    merchant = get_merchant(request)
    # 当前用户可见的工厂（merchant_admin看全部，staff看全部工厂的订单）
    factories = merchant.factories.filter(is_active=True)
    factory_ids = list(factories.values_list('id', flat=True))

    # 按生产状态分组
    base_qs = Order.objects.filter(merchant=merchant, factory_id__in=factory_ids, status='in_production')

    # 待生产：未开始
    pending_orders = base_qs.filter(production_started_at__isnull=True).order_by('created_at')
    # 生产中：已开始但未完成
    active_orders = base_qs.filter(production_started_at__isnull=False, production_completed_at__isnull=True).order_by('production_started_at')
    # 已完成：生产完成待发货
    completed_orders = base_qs.filter(production_completed_at__isnull=False).order_by('-production_completed_at')

    # 【新增】为看板订单计算工厂下载时效状态
    now = timezone.now()
    def _calc_download_sla(order):
        """计算工厂下载时效状态"""
        order.download_sla_status = 'none'
        order.download_sla_class = ''
        order.download_sla_title = ''
        if order.factory_notified_at:
            if order.factory_downloaded_at:
                order.download_sla_status = 'done'
                order.download_sla_class = 'done'
                order.download_sla_title = '已下载'
            else:
                elapsed = (now - order.factory_notified_at).total_seconds() / 60
                if elapsed > 30:
                    order.download_sla_status = 'overdue'
                    order.download_sla_class = 'overdue'
                    order.download_sla_title = f'已超{int(elapsed - 30)}分'
                elif elapsed > 15:
                    order.download_sla_status = 'warning'
                    order.download_sla_class = 'warning'
                    order.download_sla_title = f'剩{int(30 - elapsed)}分'
                else:
                    order.download_sla_status = 'normal'
                    order.download_sla_class = 'normal'
                    order.download_sla_title = '正常'
    for o in list(pending_orders) + list(active_orders) + list(completed_orders):
        _calc_download_sla(o)

    # 处理POST操作
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'start':
            order_id = request.POST.get('order_id')
            order = get_object_or_404(Order, pk=order_id, merchant=merchant)
            order.production_started_at = timezone.now()
            order.save(update_fields=['production_started_at'])
            OrderStatusLog.objects.create(
                order=order, from_status='in_production', to_status='in_production',
                operator=request.user, remark='工厂开始生产'
            )
            messages.success(request, f'订单 {order.sn} 已开始生产')
        elif action == 'complete':
            order_id = request.POST.get('order_id')
            order = get_object_or_404(Order, pk=order_id, merchant=merchant)
            order.production_completed_at = timezone.now()
            order.save(update_fields=['production_completed_at'])
            # 上传生产照片
            if request.FILES.get('photo'):
                ProductionPhoto.objects.create(
                    order=order,
                    photo_type='inspection',
                    image=request.FILES['photo'],
                    uploaded_by=request.user
                )
            OrderStatusLog.objects.create(
                order=order, from_status='in_production', to_status='in_production',
                operator=request.user, remark='工厂生产完成'
            )
            messages.success(request, f'订单 {order.sn} 已标记生产完成')
        elif action == 'ship':
            order_id = request.POST.get('order_id')
            order = get_object_or_404(Order, pk=order_id, merchant=merchant)
            ship_method = request.POST.get('ship_method', 'express')
            if request.FILES.get('photo'):
                photo_type = 'express_receipt' if ship_method == 'express' else 'delivery_photo'
                ProductionPhoto.objects.create(
                    order=order,
                    photo_type=photo_type,
                    image=request.FILES['photo'],
                    uploaded_by=request.user
                )
            order.delivery_type = 'express' if ship_method == 'express' else 'flash'
            order.tracking_company = request.POST.get('tracking_company', '')
            order.save(update_fields=['delivery_type', 'tracking_company'])
            remark = '快递发货' if ship_method == 'express' else '自行派送'
            order.shipped_at = timezone.now()
            order.save(update_fields=['shipped_at'])
            order.transition_status('shipped', operator=request.user, remark=remark)
            messages.success(request, f'订单 {order.sn} 已发货（{remark}）')
        elif action == 'extend_delivery':
            order_id = request.POST.get('order_id')
            order = get_object_or_404(Order, pk=order_id, merchant=merchant)
            new_date_str = request.POST.get('new_date')
            reason = request.POST.get('reason', '')
            if not new_date_str:
                messages.error(request, '请选择新的交货时间')
                return redirect('factory_production_board')
            from datetime import datetime
            try:
                new_date = datetime.strptime(new_date_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                messages.error(request, '交货时间格式不正确')
                return redirect('factory_production_board')
            original_date = order.delivery_date
            DeliveryExtension.objects.create(
                order=order,
                original_date=original_date,
                new_date=new_date,
                reason=reason,
                created_by=request.user
            )
            order.delivery_date = new_date
            order.save(update_fields=['delivery_date'])
            OrderStatusLog.objects.create(
                order=order, from_status='in_production', to_status='in_production',
                operator=request.user, remark=f'申请延长交货时间: {original_date} -> {new_date}，原因: {reason}'
            )
            messages.success(request, f'订单 {order.sn} 交货时间已延长至 {new_date.strftime("%m-%d %H:%M")}')
            return redirect('factory_production_board')
        elif action == 'upload_production_photo':
            order_id = request.POST.get('order_id')
            order = get_object_or_404(Order, pk=order_id, merchant=merchant)
            photo = request.FILES.get('photo')
            if photo:
                ProductionPhoto.objects.create(
                    order=order,
                    photo_type='production',
                    image=photo,
                    uploaded_by=request.user
                )
                messages.success(request, f'订单 {order.sn} 生产照片已上传')
            else:
                messages.warning(request, '未选择照片')
            return redirect('factory_production_board')
        elif action == 'batch_ship':
            shipped_count = 0
            for key, value in request.POST.items():
                if key.startswith('ship_order_'):
                    order_id = key[11:]
                    try:
                        order = Order.objects.get(pk=order_id, merchant=merchant, status='in_production', production_completed_at__isnull=False)
                    except Order.DoesNotExist:
                        continue
                    ship_method = request.POST.get(f'ship_method_{order_id}', 'express')
                    photo = request.FILES.get(f'photo_{order_id}')
                    if photo:
                        photo_type = 'express_receipt' if ship_method == 'express' else 'delivery_photo'
                        ProductionPhoto.objects.create(
                            order=order,
                            photo_type=photo_type,
                            image=photo,
                            uploaded_by=request.user
                        )
                    order.delivery_type = 'express' if ship_method == 'express' else 'flash'
                    order.tracking_company = request.POST.get(f'tracking_company_{order_id}', '')
                    order.shipped_at = timezone.now()
                    order.save(update_fields=['delivery_type', 'tracking_company', 'shipped_at'])
                    remark = '快递发货' if ship_method == 'express' else '自行派送'
                    order.transition_status('shipped', operator=request.user, remark=remark)
                    shipped_count += 1
            if shipped_count > 0:
                messages.success(request, f'已成功批量发货 {shipped_count} 个订单')
            else:
                messages.warning(request, '未选中任何订单，请勾选需要发货的订单')
        return redirect('factory_production_board')

    # 计算下次刷新时间（整点刷新，下午1点开始工作）
    next_refresh = now.replace(minute=0, second=0, microsecond=0)
    if next_refresh <= now:
        next_refresh = next_refresh.replace(hour=(next_refresh.hour + 1) % 24)
    seconds_until_refresh = int((next_refresh - now).total_seconds())

    # 即将逾期：距离交货时间不到24小时且尚未完成的订单
    from datetime import timedelta
    overdue_threshold = now + timedelta(hours=24)
    overdue_orders = base_qs.filter(
        delivery_date__isnull=False,
        delivery_date__lte=overdue_threshold,
        production_completed_at__isnull=True
    )

    # ========== 今日生产概览统计 ==========
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)

    # 今日已完成（今天生产完成的）
    today_completed = base_qs.filter(
        production_completed_at__gte=today_start,
        production_completed_at__lt=today_end
    )
    # 今日开始生产
    today_started = base_qs.filter(
        production_started_at__gte=today_start,
        production_started_at__lt=today_end
    )
    # 今日发货的
    today_shipped = merchant.orders.filter(
        status='received',
        shipped_at__gte=today_start,
        shipped_at__lt=today_end
    )

    # 计算各状态的面积（从OrderItem.area汇总）
    def calc_area(order_qs):
        from django.db.models import Sum
        total = OrderItem.objects.filter(order__in=order_qs).aggregate(s=Sum('area'))['s'] or 0
        return total

    today_stats = {
        'completed_count': today_completed.count(),
        'completed_area': calc_area(today_completed),
        'started_count': today_started.count(),
        'started_area': calc_area(today_started),
        'shipped_count': today_shipped.count(),
        'shipped_area': calc_area(today_shipped),
    }

    # ========== 工厂产能排行（按做版面积，按材料分类） ==========
    # 统计今日各工厂的做版面积（按材料分类）
    from django.db.models import Sum, Count
    factory_capacity = []
    for factory in factories:
        # 该工厂今日已完成生产的订单
        factory_orders = base_qs.filter(
            factory=factory,
            production_completed_at__gte=today_start,
            production_completed_at__lt=today_end
        )
        # 按材料分组统计面积
        material_stats = OrderItem.objects.filter(
            order__in=factory_orders
        ).values('material').annotate(
            total_area=Sum('area'),
            order_count=Count('order', distinct=True)
        ).order_by('-total_area')

        # 材料名称映射
        material_name_map = dict(OrderItem._meta.get_field('material').choices)

        factory_data = {
            'factory': factory,
            'total_orders': factory_orders.count(),
            'total_area': calc_area(factory_orders),
            'materials': [
                {
                    'name': material_name_map.get(m['material'], m['material']),
                    'area': m['total_area'] or 0,
                    'orders': m['order_count']
                }
                for m in material_stats
            ]
        }
        if factory_data['total_area'] > 0 or factory_data['total_orders'] > 0:
            factory_capacity.append(factory_data)

    # 按总面积排序
    factory_capacity.sort(key=lambda x: x['total_area'], reverse=True)

    ctx = {
        'pending_orders': pending_orders,
        'active_orders': active_orders,
        'completed_orders': completed_orders,
        'factories': factories,
        'now': now,
        'next_refresh': next_refresh,
        'seconds_until_refresh': seconds_until_refresh,
        'overdue_orders': overdue_orders,
        'stats': {
            'pending': pending_orders.count(),
            'active': active_orders.count(),
            'completed': completed_orders.count(),
            'overdue': overdue_orders.count(),
        },
        'today_stats': today_stats,
        'factory_capacity': factory_capacity,
    }
    return render(request, 'merchant/factory_production_board.html', ctx)


# ==================== 补版管理 ====================

def _can_initiate_remake(user):
    """判断当前用户是否有权限发起补版（客服/设计/管理员）"""
    if user.user_type == 'merchant_admin':
        return True
    if user.user_type == 'merchant_staff':
        profile = getattr(user, 'staff_profile', None)
        if profile and profile.is_active and profile.role:
            return profile.role.name in ('customer_service', 'designer', 'admin')
    return False


@login_required
@merchant_required
def remake_order_create(request, order_id):
    """发起补版单：复制原订单项，金额设为0，由客服/设计岗操作"""
    merchant = get_merchant(request)
    original_order = get_object_or_404(Order, pk=order_id, merchant=merchant)

    if not _can_initiate_remake(request.user):
        messages.error(request, '仅客服/设计/管理员可发起补版')
        return redirect('merchant_order_detail', order_id=order_id)

    # 补版仅对已完成或已发货的订单开放
    if original_order.status not in ('shipped', 'received', 'in_production'):
        messages.error(request, '当前订单状态不支持补版')
        return redirect('merchant_order_detail', order_id=order_id)

    if request.method == 'POST':
        item_ids = request.POST.getlist('item_ids')
        remake_reason = request.POST.get('remake_reason', '').strip()
        if not item_ids:
            messages.error(request, '请至少选择一项需要补做的产品')
            return redirect('remake_order_create', order_id=order_id)
        if not remake_reason:
            messages.error(request, '请填写补版原因')
            return redirect('remake_order_create', order_id=order_id)

        with transaction.atomic():
            # 创建补版订单
            remake_order = Order.objects.create(
                customer=original_order.customer,
                merchant=merchant,
                order_type='remake',
                original_order=original_order,
                remake_reason=remake_reason,
                remake_initiator=request.user,
                status='pending_confirm',
                is_submitted=True,
                urgent=False,
                delivery_type=original_order.delivery_type,
                delivery_address=original_order.delivery_address,
                total_amount=Decimal('0'),
            )
            # 复制选中的订单明细，金额强制为0
            for item in original_order.items.filter(id__in=item_ids):
                new_item = OrderItem.objects.create(
                    order=remake_order,
                    product_name=item.product_name,
                    material=item.material,
                    thickness=item.thickness,
                    length_mm=item.length_mm,
                    width_mm=item.width_mm,
                    quantity=item.quantity,
                    unit_price=Decimal('0'),
                    area=item.area,  # 面积按原尺寸正常计算
                    subtotal=Decimal('0'),
                    file=item.file,
                    original_file_name=item.original_file_name,
                    file_processed=item.file_processed,
                    file_standard_checked=item.file_standard_checked,
                    is_image_file=item.is_image_file,
                    plate_type=item.plate_type,
                )
            # 记录日志
            OrderStatusLog.objects.create(
                order=remake_order,
                from_status='',
                to_status='pending_confirm',
                operator=request.user,
                remark=f'补版单：关联原订单 {original_order.sn}，原因：{remake_reason}'
            )
            # 同时在原订单也记录
            OrderStatusLog.objects.create(
                order=original_order,
                from_status=original_order.status,
                to_status=original_order.status,
                operator=request.user,
                remark=f'发起补版单 {remake_order.sn}，原因：{remake_reason}'
            )
        messages.success(request, f'补版单 {remake_order.sn} 已创建，金额为0元')
        return redirect('merchant_order_detail', order_id=remake_order.id)

    return render(request, 'merchant/remake_form.html', {
        'original_order': original_order,
        'items': original_order.items.all(),
    })


# ==================== 对账单管理 ====================

@login_required
@merchant_required
@staff_permission('finance_manage')
def statement_list(request):
    """商户：对账单列表"""
    merchant = get_merchant(request)
    status_filter = request.GET.get('status')
    customer_query = request.GET.get('customer')
    statements = merchant.statements.all()
    if status_filter:
        statements = statements.filter(status=status_filter)
    if customer_query:
        statements = statements.filter(
            models.Q(customer__phone__icontains=customer_query) |
            models.Q(customer__customer_profile__company_name__icontains=customer_query)
        )
    statements = statements.order_by('-created_at')
    return render(request, 'merchant/statement_list.html', {
        'statements': statements,
        'status_choices': Statement.STATUS_CHOICES,
        'status_filter': status_filter,
    })


@login_required
@merchant_required
@staff_permission('finance_manage')
def statement_detail(request, statement_id):
    """商户：对账单详情"""
    merchant = get_merchant(request)
    statement = get_object_or_404(Statement, pk=statement_id, merchant=merchant)
    return render(request, 'merchant/statement_detail.html', {
        'statement': statement,
        'orders': statement.orders.all().order_by('created_at'),
    })


@login_required
@merchant_required
@staff_permission('finance_manage')
@transaction.atomic
def statement_generate(request):
    """商户：手动生成对账单（汇总客户未结清的已收货订单）"""
    merchant = get_merchant(request)
    if request.method == 'POST':
        order_ids_str = request.POST.get('order_ids', '')
        # 方式1：通过勾选订单批量生成
        if order_ids_str:
            order_ids = [oid.strip() for oid in order_ids_str.split(',') if oid.strip()]
            if not order_ids:
                messages.error(request, '请选择需要对账的订单')
                return redirect('merchant_orders')
            orders = Order.objects.filter(
                id__in=order_ids,
                merchant=merchant,
                status='received',
                is_settled=False,
                statement__isnull=True,
            )
            if not orders.exists():
                messages.warning(request, '选中的订单中没有可对账的订单')
                return redirect('merchant_orders')
            # 检查是否同一客户
            customers = orders.values_list('customer', flat=True).distinct()
            if len(customers) > 1:
                messages.error(request, '请选择同一客户的订单生成对账单')
                return redirect('merchant_orders')
            customer = orders.first().customer
            period_start_dt = orders.aggregate(min_date=models.Min('created_at__date'))['min_date']
            period_end_dt = orders.aggregate(max_date=models.Max('created_at__date'))['max_date']
            statement = Statement.objects.create(
                customer=customer,
                merchant=merchant,
                period_start=period_start_dt,
                period_end=period_end_dt,
            )
            for order in orders:
                order.statement = statement
                order.save(update_fields=['statement'])
            statement.update_total()
            messages.success(request, f'对账单 {statement.sn} 已生成，包含 {orders.count()} 个订单，金额 ¥{statement.total_amount}')
            return redirect('statement_detail', statement_id=statement.id)
        
        # 方式2：通过客户+周期生成
        customer_id = request.POST.get('customer_id')
        period_start = request.POST.get('period_start')
        period_end = request.POST.get('period_end')
        if not all([customer_id, period_start, period_end]):
            messages.error(request, '请填写完整的对账单信息')
            return redirect('statement_list')
        customer = get_object_or_404(User, pk=customer_id)
        from datetime import datetime
        period_start_dt = datetime.strptime(period_start, '%Y-%m-%d').date()
        period_end_dt = datetime.strptime(period_end, '%Y-%m-%d').date()
        # 查找该客户在该周期内已收货且未结清、未关联其他账单的订单
        orders = Order.objects.filter(
            customer=customer,
            merchant=merchant,
            status='received',
            is_settled=False,
            statement__isnull=True,
            created_at__date__gte=period_start_dt,
            created_at__date__lte=period_end_dt,
        )
        if not orders.exists():
            messages.warning(request, '该客户在指定周期内没有可对账的订单')
            return redirect('statement_list')
        statement = Statement.objects.create(
            customer=customer,
            merchant=merchant,
            period_start=period_start_dt,
            period_end=period_end_dt,
        )
        for order in orders:
            order.statement = statement
            order.save(update_fields=['statement'])
        statement.update_total()
        messages.success(request, f'对账单 {statement.sn} 已生成，包含 {orders.count()} 个订单，金额 ¥{statement.total_amount}')
        return redirect('statement_detail', statement_id=statement.id)
    # GET: 显示生成页面
    members = merchant.customers.filter(registration_status='approved')
    from datetime import date, timedelta
    today = date.today()
    # 默认上月周期
    if today.day >= 15:
        default_start = today.replace(day=1).strftime('%Y-%m-%d')
        default_end = today.strftime('%Y-%m-%d')
    else:
        last_month = today.replace(day=1) - timedelta(days=1)
        default_start = last_month.replace(day=1).strftime('%Y-%m-%d')
        default_end = last_month.strftime('%Y-%m-%d')
    return render(request, 'merchant/statement_generate.html', {
        'members': members,
        'default_start': default_start,
        'default_end': default_end,
    })


@login_required
@merchant_required
def statement_mark_paid(request, statement_id):
    """商户：标记对账单为已付款（客户确认后，商户核实到账标记）"""
    merchant = get_merchant(request)
    statement = get_object_or_404(Statement, pk=statement_id, merchant=merchant)
    if statement.status != 'confirmed':
        messages.error(request, '该对账单当前状态不支持标记付款')
        return redirect('statement_detail', statement_id=statement.id)
    if request.method == 'POST':
        statement.status = 'paid'
        statement.paid_at = timezone.now()
        statement.remark = request.POST.get('remark', statement.remark)
        statement.save()
        messages.success(request, f'对账单 {statement.sn} 已标记为已付款，可进行结清操作')
    return redirect('statement_detail', statement_id=statement.id)


@login_required
@merchant_required
@transaction.atomic
def statement_settle(request, statement_id):
    """商户：确认收款并结清对账单，释放客户额度"""
    merchant = get_merchant(request)
    statement = get_object_or_404(Statement, pk=statement_id, merchant=merchant)
    if statement.status not in ('pending', 'confirmed', 'paid'):
        messages.error(request, '该对账单当前状态不支持结清操作')
        return redirect('statement_detail', statement_id=statement.id)
    if request.method == 'POST':
        statement.status = 'settled'
        statement.settled_at = timezone.now()
        statement.settled_by = request.user
        statement.remark = request.POST.get('remark', statement.remark)
        statement.save()
        # 标记账单下所有订单为已结清
        total_released = Decimal('0')
        for order in statement.orders.all():
            if not order.is_settled:
                order.is_settled = True
                order.save(update_fields=['is_settled'])
                total_released += order.total_amount
        # 释放客户额度
        profile = statement.customer.customer_profile
        profile.credit_used -= total_released
        if profile.credit_used < 0:
            profile.credit_used = 0
        profile.save(update_fields=['credit_used'])
        messages.success(
            request,
            f'对账单 {statement.sn} 已结清，释放客户额度 ¥{total_released}，'
            f'客户当前已用额度 ¥{profile.credit_used} / ¥{profile.credit_limit}'
        )
    return redirect('statement_detail', statement_id=statement.id)


# ==================== 对账单导出Excel ====================

def _build_statement_excel(statement):
    """生成对账单Excel工作簿，返回HttpResponse"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from decimal import Decimal

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '对账单'

    # 样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    header_font = Font(name='微软雅黑', size=11, bold=True)
    normal_font = Font(name='微软雅黑', size=10)
    small_font = Font(name='微软雅黑', size=9, color='666666')
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 第1行：商户标题
    ws.merge_cells('A1:L1')
    ws['A1'] = f'{statement.merchant.name} — 对账单'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30

    # 第2行：空行
    ws.row_dimensions[2].height = 8

    # 第3行：单号
    ws.merge_cells('A3:F3')
    ws['A3'] = f'对账单号：{statement.sn}'
    ws['A3'].font = normal_font
    ws['A3'].alignment = left_align

    ws.merge_cells('G3:L3')
    ws['G3'] = f'账单周期：{statement.period_start.strftime("%Y年%m月%d日")} 至 {statement.period_end.strftime("%Y年%m月%d日")}'
    ws['G3'].font = normal_font
    ws['G3'].alignment = Alignment(horizontal='right', vertical='center')

    # 第4行：客户信息
    ws.merge_cells('A4:F4')
    profile = statement.customer.customer_profile
    ws['A4'] = f'客户名称：{profile.company_name or profile.real_name or statement.customer.phone}'
    ws['A4'].font = normal_font
    ws['A4'].alignment = left_align

    ws.merge_cells('G4:L4')
    ws['G4'] = f'联系方式：{statement.customer.phone}'
    ws['G4'].font = normal_font
    ws['G4'].alignment = Alignment(horizontal='right', vertical='center')

    # 第5行：空行
    ws.row_dimensions[5].height = 8

    # 第6行：表头
    headers = ['序号', '订单号', '产品名称', '材质', '厚度', '长度(mm)', '宽度(mm)', '数量', '单价(元/cm²)', '金额(元)', '文件名称']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    ws.row_dimensions[6].height = 25

    # 数据行：按订单项展开
    row_idx = 7
    seq = 1
    total_amount = Decimal('0')

    for order in statement.orders.all().order_by('created_at'):
        items = order.items.all()
        if not items:
            continue
        for item in items:
            data = [
                seq,
                order.sn,
                item.get_product_name_display(),
                item.get_material_display(),
                item.thickness,
                float(item.length_mm),
                float(item.width_mm),
                item.quantity,
                float(item.unit_price),
                float(item.subtotal),
                item.file.name.split('/')[-1] if item.file else ''
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = normal_font
                cell.border = border
                if col in (1, 2, 4, 5, 8):
                    cell.alignment = center_align
                elif col in (6, 7, 9, 10):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = left_align
            row_idx += 1
            seq += 1
            total_amount += item.subtotal

    # 如果订单有紧急费，单独列一行
    for order in statement.orders.all().order_by('created_at'):
        if order.urgent:
            urgent_fee = (order.total_amount - sum(i.subtotal for i in order.items.all())).quantize(Decimal('0.01'))
            if urgent_fee > 0:
                data = [
                    seq,
                    order.sn,
                    '加急费',
                    '-',
                    '-',
                    '-',
                    '-',
                    '-',
                    '-',
                    float(urgent_fee),
                    ''
                ]
                for col, val in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.font = normal_font
                    cell.border = border
                    if col in (1, 2, 4, 5, 8):
                        cell.alignment = center_align
                    elif col in (6, 7, 9, 10):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = left_align
                row_idx += 1
                seq += 1
                total_amount += urgent_fee

    # 合计行
    ws.merge_cells(f'A{row_idx}:I{row_idx}')
    ws.cell(row=row_idx, column=1, value=f'合计金额（大写）：{_num_to_chinese(total_amount)}')
    ws.cell(row=row_idx, column=1).font = Font(name='微软雅黑', size=11, bold=True)
    ws.cell(row=row_idx, column=1).alignment = left_align
    ws.cell(row=row_idx, column=1).border = border
    for c in range(2, 10):
        ws.cell(row=row_idx, column=c).border = border

    ws.cell(row=row_idx, column=10, value=f'¥{total_amount}')
    ws.cell(row=row_idx, column=10).font = Font(name='微软雅黑', size=11, bold=True, color='C00000')
    ws.cell(row=row_idx, column=10).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=row_idx, column=10).border = border
    ws.cell(row=row_idx, column=11).border = border
    ws.row_dimensions[row_idx].height = 28
    row_idx += 1

    # 底部备注
    ws.merge_cells(f'A{row_idx}:L{row_idx}')
    ws.cell(row=row_idx, column=1, value='注：请收货后及时核对产品，若有不符，请于3天内与本公司联系。')
    ws.cell(row=row_idx, column=1).font = small_font
    ws.cell(row=row_idx, column=1).alignment = left_align
    row_idx += 1

    ws.merge_cells(f'A{row_idx}:F{row_idx}')
    ws.cell(row=row_idx, column=1, value='')
    ws.cell(row=row_idx, column=1).font = normal_font
    ws.cell(row=row_idx, column=1).alignment = left_align

    ws.merge_cells(f'G{row_idx}:L{row_idx}')
    ws.cell(row=row_idx, column=7, value='签收人：________________')
    ws.cell(row=row_idx, column=7).font = normal_font
    ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='right', vertical='center')

    # 列宽
    col_widths = [6, 18, 22, 10, 8, 12, 12, 8, 14, 12, 28]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 保存到内存
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _num_to_chinese(num):
    """数字金额转中文大写（简化版）"""
    from decimal import Decimal
    num = Decimal(str(num))
    integer_part = int(num)
    decimal_part = (num - integer_part) * 100
    jiao = int(decimal_part // 10)
    fen = int(decimal_part % 10)

    digit_map = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    unit_map = ['', '拾', '佰', '仟']
    big_unit = ['', '万', '亿']

    def int_to_chinese(n):
        if n == 0:
            return '零'
        s = str(n)
        res = ''
        zero_flag = False
        for i, ch in enumerate(s):
            d = int(ch)
            pos = len(s) - i - 1
            if d == 0:
                if not zero_flag and pos % 4 == 0 and res:
                    pass
                elif not zero_flag:
                    zero_flag = True
            else:
                if zero_flag:
                    res += '零'
                    zero_flag = False
                res += digit_map[d] + unit_map[pos % 4]
            if pos % 4 == 0 and pos > 0 and any(int(c) for c in s[max(0, i-3):i+1]):
                res += big_unit[pos // 4]
        return res

    result = int_to_chinese(integer_part) + '元'
    if jiao == 0 and fen == 0:
        result += '整'
    else:
        if jiao > 0:
            result += digit_map[jiao] + '角'
        if fen > 0:
            result += digit_map[fen] + '分'
    return result


@login_required
@merchant_required
def statement_export(request, statement_id):
    """商户：导出对账单Excel"""
    merchant = get_merchant(request)
    statement = get_object_or_404(Statement, pk=statement_id, merchant=merchant)
    output = _build_statement_excel(statement)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"对账单_{statement.sn}_{statement.customer.phone}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



# ==================== 拼版批次管理（跨订单拼版）====================

@login_required
@merchant_required
@staff_permission('design_layout')
def plate_batch_list(request):
    """拼版批次列表页"""
    merchant = get_merchant(request)
    batches = PlateBatch.objects.filter(merchant=merchant).prefetch_related('items', 'items__order', 'items__order_item')

    status_filter = request.GET.get('status', '')
    if status_filter:
        batches = batches.filter(status=status_filter)

    batches = batches.order_by('-created_at')

    # 统计
    stats = {
        'total': batches.count(),
        'auto_generated': batches.filter(status='auto_generated').count(),
        'confirmed': batches.filter(status='confirmed').count(),
        'in_production': batches.filter(status='in_production').count(),
    }

    # 为每个批次预计算客户数（避免模板中执行复杂查询）
    batch_list = []
    for batch in batches:
        customer_ids = set()
        for bi in batch.items.all():
            if bi.order and bi.order.customer_id:
                customer_ids.add(bi.order.customer_id)
        batch.customer_count = len(customer_ids)
        batch_list.append(batch)

    return render(request, 'merchant/plate_batch_list.html', {
        'batches': batch_list,
        'stats': stats,
        'status_filter': status_filter,
    })


@login_required
@merchant_required
def plate_batch_generate(request):
    """触发自动跨订单拼版"""
    if request.method != 'POST':
        return redirect('plate_batch_list')

    merchant = get_merchant(request)
    from utils.plate_batch import auto_generate_plate_batches

    try:
        results = auto_generate_plate_batches(merchant, algorithm='maxrects')
        if results:
            messages.success(request, f'成功生成 {len(results)} 个拼版批次')
        else:
            messages.info(request, '当前没有待拼版的订单')
    except Exception as e:
        messages.error(request, f'拼版生成失败: {str(e)}')

    return redirect('plate_batch_list')


@login_required
@merchant_required
@staff_permission('design_layout')
def plate_batch_detail(request, batch_id):
    """拼版批次详情/工作台（支持拖拽微调）"""
    merchant = get_merchant(request)
    batch = get_object_or_404(PlateBatch, pk=batch_id, merchant=merchant)

    # 算法选择
    ALGORITHMS = [
        ('maxrects', 'MaxRects（推荐·利用率最高）'),
        ('guillotine', 'Guillotine（平衡型）'),
        ('skyline', 'Skyline（快速型）'),
    ]
    algorithm = request.GET.get('algorithm', 'maxrects')
    if algorithm not in [a[0] for a in ALGORITHMS]:
        algorithm = 'maxrects'

    # 重新生成建议（如果点击了重新计算）
    suggestion = None
    if request.GET.get('recalculate') == '1' and batch.status == 'auto_generated':
        from utils.plate_type_rules import get_spacing_mm
        from utils.plate_batch import build_rects_from_items, pack_rects_into_plates
        batch_items = [bi.order_item for bi in batch.items.all()]
        if batch_items:
            spacing = get_spacing_mm(batch.thickness, batch.thickness, batch.material, batch.material)
            rects = build_rects_from_items(batch_items, spacing_mm=spacing)
            plates = pack_rects_into_plates(rects, algorithm=algorithm)
            if plates:
                suggestion = plates[0]

    # 解析已保存的布局数据
    layout_data_parsed = {}
    if batch.layout_data:
        try:
            layout_data_parsed = json.loads(batch.layout_data)
        except Exception:
            pass

    # 画布缩放（增大画布让预览图更清晰）
    CANVAS_WIDTH = 1400
    CANVAS_HEIGHT = 1200
    scaled_rectangles = []
    scaled_plate_width = 0
    scaled_plate_height = 0
    scale = 1.0

    # 预加载文件信息用于生成预览图（优先使用制版文件）
    batch_items_qs = batch.items.select_related('order_item').all()
    item_file_map = {}
    for bi in batch_items_qs:
        if bi.order_item_id:
            oi = bi.order_item
            # 优先使用 plate_file，其次使用客户源文件 file
            if hasattr(oi, 'plate_file') and oi.plate_file:
                item_file_map[str(bi.order_item_id)] = oi.plate_file.name
            elif oi.file:
                item_file_map[str(bi.order_item_id)] = oi.file.name

    display_data = layout_data_parsed or suggestion
    if display_data:
        pw = float(display_data.get('plate_width', batch.plate_width))
        ph = float(display_data.get('plate_height', batch.plate_height))
        scale_w = CANVAS_WIDTH / pw
        scale_h = CANVAS_HEIGHT / ph
        scale = min(scale_w, scale_h)
        scaled_plate_width = round(pw * scale, 1)
        scaled_plate_height = round(ph * scale, 1)

        order_hues = {}
        hue_idx = 0
        for rect in display_data.get('rectangles', []):
            order_sn = rect.get('order_sn', '')
            if order_sn not in order_hues:
                order_hues[order_sn] = (hue_idx * 47) % 360
                hue_idx += 1
            rotation = int(rect.get('rotation', 0)) % 360
            rw = round(float(rect.get('width', 0)) * scale, 1)
            rh = round(float(rect.get('height', 0)) * scale, 1)
            # 若旋转90/270度，画布上交换宽高显示
            if rotation in (90, 270):
                rw, rh = rh, rw

            # 生成带版类视觉效果的预览图
            preview_url = None
            rid = rect.get('id', '')
            order_item_id = rid.split('_')[0] if '_' in rid else rid
            file_name = item_file_map.get(order_item_id)
            if file_name:
                preview_rel = f"plate_previews/{batch.id.hex}/{order_item_id}.png"
                preview_path = os.path.join(settings.MEDIA_ROOT, preview_rel)
                if not os.path.exists(preview_path):
                    # 获取对应 OrderItem 的产品名称以确定版类效果
                    order_item = None
                    for bi in batch_items_qs:
                        if str(bi.order_item_id) == order_item_id:
                            order_item = bi.order_item
                            break
                    if order_item:
                        from utils.plate_preview_effects import generate_effect_preview
                        generate_effect_preview(
                            file_name, preview_rel,
                            product_name=order_item.product_name,
                            plate_type_key=getattr(order_item, 'plate_type', None),
                            dpi=150
                        )
                    else:
                        from utils.pdf_processor import generate_pdf_preview
                        generate_pdf_preview(file_name, preview_rel, dpi=72)
                if os.path.exists(preview_path):
                    preview_url = settings.MEDIA_URL + preview_rel

            scaled_rectangles.append({
                'id': rid,
                'x': round(float(rect.get('x', 0)) * scale, 1),
                'y': round(float(rect.get('y', 0)) * scale, 1),
                'width': rw,
                'height': rh,
                'mm_width': float(rect.get('width', 0)),
                'mm_height': float(rect.get('height', 0)),
                'label': rect.get('label', ''),
                'order_sn': order_sn,
                'customer_phone': rect.get('customer_phone', ''),
                'color_hue': order_hues[order_sn],
                'rotation': rotation,
                'preview_url': preview_url,
            })

    if request.method == 'POST':
        action = request.POST.get('action', 'save')

        if action == 'save':
            data = json.loads(request.POST.get('layout_data', '{}'))
            note = request.POST.get('designer_note', '')
            batch.layout_data = json.dumps(data, ensure_ascii=False)
            batch.designer_note = note
            batch.designer = request.user
            batch.save()
            messages.success(request, '拼版布局已保存')
            return redirect('plate_batch_detail', batch_id=batch.id)

        elif action == 'confirm':
            data = json.loads(request.POST.get('layout_data', '{}'))
            note = request.POST.get('designer_note', '')
            batch.layout_data = json.dumps(data, ensure_ascii=False)
            batch.designer_note = note
            batch.designer = request.user
            batch.status = 'confirmed'
            default_factory = merchant.factories.filter(is_active=True).first()
            if default_factory:
                batch.factory = default_factory
            batch.save()

            # === 阶段4：确认时重新生成高清生产PDF和效果图 ===
            try:
                from utils.plate_pdf import generate_plate_production_pdf
                pdf_rel_path, pdf_url = generate_plate_production_pdf(batch, use_plate_file=True)
                if pdf_rel_path:
                    batch.production_pdf = pdf_rel_path
                    batch.save(update_fields=['production_pdf'])
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'[生产PDF生成失败] batch={batch.id.hex[:8]} error={e}')

            try:
                from utils.plate_batch import generate_plate_image
                # 使用当前 layout_data 重新生成高清效果图（300 DPI）
                plate_result = {
                    'spec': {
                        'width': batch.plate_width,
                        'height': batch.plate_height,
                        'name': batch.plate_spec_name,
                    },
                    'placed': [],
                    'usage_rate': batch.usage_rate or 0,
                }
                for rect in data.get('rectangles', []):
                    # 找到对应的 OrderItem
                    rid = rect.get('id', '')
                    order_item_id = rid.rsplit('_', 1)[0] if '_' in rid else rid
                    item = None
                    for bi in batch.items.select_related('order_item').all():
                        if str(bi.order_item_id) == order_item_id:
                            item = bi.order_item
                            break
                    plate_result['placed'].append({
                        'id': rid,
                        'x': rect.get('x', 0),
                        'y': rect.get('y', 0),
                        'orig_width': rect.get('width', 0),
                        'orig_height': rect.get('height', 0),
                        'label': rect.get('label', ''),
                        'order_sn': rect.get('order_sn', ''),
                        'customer_phone': rect.get('customer_phone', ''),
                        'original_item': item,
                    })
                img_rel_path, img_url = generate_plate_image(plate_result, dpi=300, use_plate_file=True)
                if img_rel_path:
                    batch.layout_image = img_rel_path
                    batch.save(update_fields=['layout_image'])
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'[高清效果图生成失败] batch={batch.id.hex[:8]} error={e}')
            # === 阶段4结束 ===

            # 更新关联订单状态为 confirmed 并下发工厂
            affected_orders = set()
            for bi in batch.items.all():
                bi.order_item.order.plate_status = 'confirmed'
                bi.order_item.order.factory = default_factory
                bi.order_item.order.save(update_fields=['plate_status', 'factory'])
                affected_orders.add(bi.order_item.order)

            for order in affected_orders:
                order.transition_status('in_production', operator=request.user,
                                        remark=f'拼版批次 {batch.id.hex[:8]} 已确认，自动下发工厂生产')

            messages.success(request, f'拼版批次已确认，共影响 {len(affected_orders)} 个订单，已自动下发工厂')
            return redirect('plate_batch_list')

        elif action == 'reject':
            reason = request.POST.get('reject_reason', '')
            batch.status = 'rejected'
            batch.save(update_fields=['status'])

            affected_orders = set()
            for bi in batch.items.all():
                bi.order_item.order.plate_status = 'rejected'
                bi.order_item.order.save(update_fields=['plate_status'])
                affected_orders.add(bi.order_item.order)

            for order in affected_orders:
                OrderStatusLog.objects.create(
                    order=order, from_status=order.status, to_status=order.status,
                    operator=request.user, remark=f'拼版批次被驳回: {reason}'
                )

            messages.warning(request, f'拼版批次已驳回，原因：{reason}')
            return redirect('plate_batch_list')

        elif action == 'upload_manual':
            # 【新增】手动上传拼版文件
            if request.FILES.get('manual_plate_file'):
                uploaded_file = request.FILES['manual_plate_file']
                ext = os.path.splitext(uploaded_file.name)[1].lower()
                if ext != '.pdf':
                    messages.error(request, '仅支持PDF格式文件')
                    return redirect('plate_batch_detail', batch_id=batch_id)
                
                # 保存文件
                filename = f"plate_layouts/{batch.id.hex}_manual{ext}"
                path = default_storage.save(filename, uploaded_file)
                
                # 更新批次
                batch.production_pdf = path
                batch.designer_note = request.POST.get('manual_note', '')
                batch.designer = request.user
                batch.status = 'confirmed'
                default_factory = merchant.factories.filter(is_active=True).first()
                if default_factory:
                    batch.factory = default_factory
                batch.save()
                
                # 更新关联订单状态
                affected_orders = set()
                for bi in batch.items.all():
                    bi.order_item.order.plate_status = 'confirmed'
                    bi.order_item.order.factory = default_factory
                    bi.order_item.order.save(update_fields=['plate_status', 'factory'])
                    affected_orders.add(bi.order_item.order)
                
                for order in affected_orders:
                    order.transition_status('in_production', operator=request.user,
                                            remark=f'拼版批次 {batch.id.hex[:8]} 已手动上传确认，下发工厂生产')
                
                messages.success(request, f'拼版文件已手动上传并确认，共影响 {len(affected_orders)} 个订单')
                return redirect('plate_batch_list')
            else:
                messages.error(request, '请选择要上传的文件')
                return redirect('plate_batch_detail', batch_id=batch_id)

    # 为侧边栏订单列表生成颜色 + 计算客户数
    sidebar_items = []
    sidebar_hue_idx = 0
    seen_orders = set()
    customer_ids = set()
    for bi in batch.items.select_related('order', 'order_item').all():
        if bi.order_id not in seen_orders:
            seen_orders.add(bi.order_id)
            sidebar_items.append({
                'order_sn': bi.order.sn,
                'length_mm': bi.order_item.length_mm,
                'width_mm': bi.order_item.width_mm,
                'color_hue': (sidebar_hue_idx * 47) % 360,
            })
            sidebar_hue_idx += 1
        if bi.order and bi.order.customer_id:
            customer_ids.add(bi.order.customer_id)
    batch.customer_count = len(customer_ids)

    return render(request, 'merchant/plate_batch_detail.html', {
        'batch': batch,
        'suggestion': suggestion,
        'layout_data_parsed': layout_data_parsed,
        'scaled_rectangles': scaled_rectangles,
        'scaled_plate_width': scaled_plate_width,
        'scaled_plate_height': scaled_plate_height,
        'algorithm': algorithm,
        'algorithms': ALGORITHMS,
        'scale': scale,
        'sidebar_items': sidebar_items,
    })


@login_required
@merchant_required
def plate_batch_confirm(request, batch_id):
    """确认拼版批次（快捷操作）"""
    if request.method != 'POST':
        return redirect('plate_batch_detail', batch_id=batch_id)

    merchant = get_merchant(request)
    batch = get_object_or_404(PlateBatch, pk=batch_id, merchant=merchant)

    if batch.status != 'auto_generated':
        messages.warning(request, '该拼版批次状态不允许确认')
        return redirect('plate_batch_list')

    default_factory = merchant.factories.filter(is_active=True).first()
    batch.status = 'confirmed'
    batch.designer = request.user
    if default_factory:
        batch.factory = default_factory
    batch.save()

    affected_orders = set()
    for bi in batch.items.all():
        order = bi.order_item.order
        order.plate_status = 'confirmed'
        if default_factory:
            order.factory = default_factory
        order.save(update_fields=['plate_status', 'factory'])
        affected_orders.add(order)

    for order in affected_orders:
        order.transition_status('in_production', operator=request.user,
                                remark=f'拼版批次 {batch.id.hex[:8]} 已确认，自动下发工厂生产')

    messages.success(request, f'拼版批次已确认，共影响 {len(affected_orders)} 个订单')
    return redirect('plate_batch_list')


@login_required
@merchant_required
def plate_batch_reject(request, batch_id):
    """驳回拼版批次"""
    if request.method != 'POST':
        return redirect('plate_batch_detail', batch_id=batch_id)

    merchant = get_merchant(request)
    batch = get_object_or_404(PlateBatch, pk=batch_id, merchant=merchant)
    reason = request.POST.get('reject_reason', '')

    batch.status = 'rejected'
    batch.save(update_fields=['status'])

    affected_orders = set()
    for bi in batch.items.all():
        order = bi.order_item.order
        order.plate_status = 'rejected'
        order.save(update_fields=['plate_status'])
        affected_orders.add(order)

    for order in affected_orders:
        OrderStatusLog.objects.create(
            order=order, from_status=order.status, to_status=order.status,
            operator=request.user, remark=f'拼版批次被驳回: {reason}'
        )

    messages.warning(request, f'拼版批次已驳回')
    return redirect('plate_batch_list')


@login_required
@merchant_required
def plate_batch_update_layout(request, batch_id):
    """AJAX：更新拼版布局坐标（拖拽微调后保存）"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '仅支持POST'})

    merchant = get_merchant(request)
    batch = get_object_or_404(PlateBatch, pk=batch_id, merchant=merchant)

    try:
        data = json.loads(request.body)
        rectangles = data.get('rectangles', [])

        # 更新 layout_data 中的坐标
        layout_data = {}
        if batch.layout_data:
            try:
                layout_data = json.loads(batch.layout_data)
            except Exception:
                pass

        existing_rects = {r['id']: r for r in layout_data.get('rectangles', [])}
        for rect in rectangles:
            rid = rect.get('id')
            if rid in existing_rects:
                existing_rects[rid]['x'] = rect.get('x', existing_rects[rid]['x'])
                existing_rects[rid]['y'] = rect.get('y', existing_rects[rid]['y'])
                if 'width' in rect:
                    existing_rects[rid]['width'] = rect['width']
                if 'height' in rect:
                    existing_rects[rid]['height'] = rect['height']
                if 'rotation' in rect:
                    existing_rects[rid]['rotation'] = rect['rotation']

        layout_data['rectangles'] = list(existing_rects.values())
        batch.layout_data = json.dumps(layout_data, ensure_ascii=False)
        batch.designer = request.user
        batch.save()

        return JsonResponse({'success': True})
    except Exception:
        logger.exception('拼版布局保存失败')
        return JsonResponse({'success': False, 'error': '保存失败，请稍后重试'})


@login_required
@merchant_required
def download_order_file(request, order_id, item_id):
    """
    安全下载订单文件
    验证当前用户是订单所属商家的员工后才允许下载
    支持 OSS 存储和本地存储两种模式
    """
    try:
        merchant = request.user.managed_merchant if request.user.user_type == 'merchant_admin' else request.user.staff_profile.merchant
    except AttributeError:
        messages.error(request, '无权访问')
        return redirect('merchant_dashboard')

    order = get_object_or_404(Order, id=order_id, merchant=merchant)
    item = get_object_or_404(OrderItem, id=item_id, order=order)

    if not item.file:
        messages.error(request, '该订单没有上传文件')
        return redirect('merchant_order_detail', order_id=order.id)

    from urllib.parse import quote
    filename = item.original_file_name or os.path.basename(item.file.name) or 'download'

    # 检测是否使用 OSS 存储
    storage_class_name = item.file.storage.__class__.__name__
    is_oss = 'OSS' in storage_class_name

    # OSS 或本地存储：统一通过服务器读取文件内容返回
    # 避免浏览器直接访问 OSS 内网 URL
    try:
        # 【新增】记录工厂下载时间（如果订单已在生产中）
        from django.utils import timezone
        if order.status == 'in_production' and order.factory_notified_at and not order.factory_downloaded_at:
            order.factory_downloaded_at = timezone.now()
            order.save(update_fields=['factory_downloaded_at'])
        
        with item.file.open('rb') as f:
            # 根据文件扩展名设置正确的 Content-Type
            ext = os.path.splitext(filename)[1].lower()
            content_type_map = {
                '.pdf': 'application/pdf',
                '.ai': 'application/postscript',
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.png': 'image/png',
                '.gif': 'image/gif',
                '.zip': 'application/zip',
            }
            content_type = content_type_map.get(ext, 'application/octet-stream')
            
            response = HttpResponse(f.read(), content_type=content_type)
            # 使用安全的文件名编码
            safe_filename = quote(filename)
            response['Content-Disposition'] = f"attachment; filename*=UTF-8''{safe_filename}"
            # 添加安全头，避免浏览器拦截
            response['X-Content-Type-Options'] = 'nosniff'
            response['Cache-Control'] = 'private, max-age=0'
            return response
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"下载文件失败: {e}", exc_info=True)
        messages.error(request, '文件不存在或已被删除')
        return redirect('merchant_order_detail', order_id=order.id)


@login_required
@merchant_required
def download_plate_batch_file(request, batch_id, field_name):
    """
    安全下载拼版批次文件（production_pdf 或 layout_image）
    """
    ALLOWED_FIELDS = {'production_pdf', 'layout_image'}
    if field_name not in ALLOWED_FIELDS:
        messages.error(request, '无效的文件类型')
        return redirect('merchant_dashboard')

    merchant = get_merchant(request)
    if not merchant:
        messages.error(request, '无权访问')
        return redirect('merchant_dashboard')

    batch = get_object_or_404(PlateBatch, id=batch_id, merchant=merchant)
    file_field = getattr(batch, field_name, None)
    if not file_field:
        messages.error(request, '文件不存在')
        return redirect('plate_batch_detail', batch_id=batch.id)

    try:
        with file_field.open('rb') as f:
            # 根据文件类型设置content_type
            ext = os.path.splitext(file_field.name)[1].lower()
            content_type = 'application/pdf' if ext == '.pdf' else 'image/png' if ext in ('.png', '.jpg', '.jpeg') else 'application/octet-stream'
            response = HttpResponse(f.read(), content_type=content_type)
            from urllib.parse import quote
            filename = os.path.basename(file_field.name)
            response['Content-Disposition'] = f"inline; filename*=UTF-8''{quote(filename)}"
            return response
    except Exception:
        messages.error(request, '文件不存在或已被删除')
        return redirect('merchant_dashboard')


# ==================== 制版文件上传功能（阶段1）====================

@login_required
@merchant_required
def pending_plate_orders(request):
    """
    待上传制版文件的订单列表
    显示状态为 pending_confirm / design_confirmed / paid 的订单项
    且没有上传过制版文件的
    """
    try:
        merchant = request.user.managed_merchant if request.user.user_type == 'merchant_admin' else request.user.staff_profile.merchant
    except AttributeError:
        messages.error(request, '无权访问')
        return redirect('merchant_dashboard')

    # 获取筛选条件
    status_filter = request.GET.get('status', '')
    
    # 基础查询：已提交的订单，且状态在生产前
    orders = merchant.orders.filter(
        is_submitted=True,
        status__in=['pending_confirm', 'design_confirmed', 'paid']
    ).select_related('customer__customer_profile').prefetch_related('items').order_by('-created_at')
    
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    # 收集需要显示的数据
    order_items_data = []
    for order in orders:
        for item in order.items.all():
            # 只显示有客户文件但还没有制版文件的订单项
            if item.file and not item.plate_file:
                order_items_data.append({
                    'order': order,
                    'item': item,
                })
    
    return render(request, 'merchant/pending_plate_orders.html', {
        'order_items_data': order_items_data,
        'status_choices': Order.STATUS_CHOICES,
        'status_filter': status_filter,
    })


@login_required
@merchant_required
def upload_plate_file(request):
    """
    AJAX接口：上传制版文件
    接收 order_item_id + 文件，保存到 OrderItem.plate_file
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '请使用POST请求'})
    
    try:
        merchant = request.user.managed_merchant if request.user.user_type == 'merchant_admin' else request.user.staff_profile.merchant
    except AttributeError:
        return JsonResponse({'success': False, 'error': '无权访问'})
    
    item_id = request.POST.get('item_id')
    if not item_id:
        return JsonResponse({'success': False, 'error': '缺少订单项ID'})
    
    # 获取订单项并验证权限
    item = get_object_or_404(OrderItem, id=item_id, order__merchant=merchant)
    
    # 检查是否有文件上传
    if 'plate_file' not in request.FILES:
        return JsonResponse({'success': False, 'error': '请选择要上传的文件'})
    
    uploaded_file = request.FILES['plate_file']
    ext = os.path.splitext(uploaded_file.name)[1].lower()
    
    # 验证文件格式
    if ext not in ('.pdf', '.ai'):
        return JsonResponse({'success': False, 'error': '仅支持PDF或AI格式文件'})
    
    try:
        # 保存文件
        filename = f"plate_files/{item.order.customer.id}/{uuid.uuid4().hex}{ext}"
        path = default_storage.save(filename, uploaded_file)
        
        # 更新订单项
        item.plate_file = path
        item.plate_file_uploaded_at = timezone.now()
        item.plate_file_uploaded_by = request.user
        item.save(update_fields=['plate_file', 'plate_file_uploaded_at', 'plate_file_uploaded_by'])
        
        return JsonResponse({
            'success': True,
            'message': '制版文件上传成功',
            'file_name': uploaded_file.name,
            'file_url': default_storage.url(path) if hasattr(default_storage, 'url') else path,
        })
    except Exception:
        logger.exception('制版文件上传失败')
        return JsonResponse({'success': False, 'error': '上传失败，请稍后重试'})


@login_required
@merchant_required
def download_plate_file(request, order_id, item_id):
    """
    下载制版文件
    """
    try:
        merchant = request.user.managed_merchant if request.user.user_type == 'merchant_admin' else request.user.staff_profile.merchant
    except AttributeError:
        messages.error(request, '无权访问')
        return redirect('merchant_dashboard')
    
    order = get_object_or_404(Order, id=order_id, merchant=merchant)
    item = get_object_or_404(OrderItem, id=item_id, order=order)
    
    if not item.plate_file:
        messages.error(request, '该订单项没有上传制版文件')
        return redirect('pending_plate_orders')
    
    try:
        # 【新增】记录工厂下载时间
        from django.utils import timezone
        if order.factory_notified_at and not order.factory_downloaded_at:
            order.factory_downloaded_at = timezone.now()
            order.save(update_fields=['factory_downloaded_at'])
        
        with item.plate_file.open('rb') as f:
            response = HttpResponse(f.read(), content_type='application/octet-stream')
            from urllib.parse import quote
            filename = os.path.basename(item.plate_file.name)
            response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
            return response
    except Exception:
        messages.error(request, '文件不存在或已被删除')
        return redirect('pending_plate_orders')


# ==================== 阶段2：订单选择拼版 ====================

@login_required
@merchant_required
@staff_permission('design_layout')
def pending_layout_orders(request):
    """
    待拼版页面（合并原拼版工具功能）
    上半部分：显示已有的拼版批次列表
    下半部分：显示可拼版的订单项（按 product_name, material, thickness 分组）
    """
    try:
        merchant = request.user.managed_merchant if request.user.user_type == 'merchant_admin' else request.user.staff_profile.merchant
    except AttributeError:
        messages.error(request, '无权访问')
        return redirect('merchant_dashboard')
    
    # ===== 上半部分：已有拼版批次 =====
    from apps.orders.models import PlateBatch
    status_filter = request.GET.get('status', '')
    batches_qs = PlateBatch.objects.filter(merchant=merchant).prefetch_related('items', 'items__order', 'items__order_item')
    if status_filter:
        batches_qs = batches_qs.filter(status=status_filter)
    batches_qs = batches_qs.order_by('-created_at')
    
    # 统计
    stats = {
        'total': PlateBatch.objects.filter(merchant=merchant).count(),
        'auto_generated': PlateBatch.objects.filter(merchant=merchant, status='auto_generated').count(),
        'confirmed': PlateBatch.objects.filter(merchant=merchant, status='confirmed').count(),
        'in_production': PlateBatch.objects.filter(merchant=merchant, status='in_production').count(),
    }
    
    # 为每个批次预计算客户数
    batch_list = []
    for batch in batches_qs:
        customer_ids = set()
        for bi in batch.items.all():
            if bi.order and bi.order.customer_id:
                customer_ids.add(bi.order.customer_id)
        batch.customer_count = len(customer_ids)
        batch_list.append(batch)
    
    # ===== 下半部分：可拼版订单项 =====
    items = OrderItem.objects.filter(
        order__merchant=merchant,
        order__status__in=['design_confirmed', 'paid'],
        order__is_submitted=True,
        plate_file__isnull=False,
    ).exclude(
        plate_batch__status='confirmed'
    ).select_related('order', 'order__customer__customer_profile').order_by('order__created_at')
    
    # 【新增】为每个item准备预览图URL
    from utils.pdf_processor import generate_pdf_preview
    for item in items:
        if item.preview_image:
            item.preview_url = item.preview_image.url
        elif item.plate_file:
            # 尝试生成预览图（生成后自动上传到OSS/本地存储）
            try:
                preview_filename = f"previews/{item.id}_plate.png"
                preview_url = generate_pdf_preview(item.plate_file.name, preview_filename, dpi=72)
                item.preview_url = preview_url
            except Exception:
                logger.exception("预览图生成失败 item=%s", item.id)
                item.preview_url = None
        else:
            item.preview_url = None
    
    # 按 (product_name, material, thickness) 分组
    groups = {}
    for item in items:
        key = (item.product_name, item.material, item.thickness)
        group_label = f"{item.get_product_name_display()} | {item.get_material_display()} | {item.thickness}mm"
        if key not in groups:
            groups[key] = {
                'label': group_label,
                'key': '_'.join(key),
                'items': [],
            }
        groups[key]['items'].append(item)
    
    return render(request, 'merchant/pending_layout_orders.html', {
        'batches': batch_list,
        'stats': stats,
        'status_filter': status_filter,
        'groups': groups,
        'total_items': items.count(),
    })


@login_required
@merchant_required
@staff_permission('design_layout')
def create_plate_batch(request):
    """
    从勾选的订单项创建拼版批次
    POST参数：
        - item_ids[]: 勾选的OrderItem ID列表
        - algorithm: 拼版算法（默认maxrects）
    """
    if request.method != 'POST':
        return redirect('pending_layout_orders')

    merchant = get_merchant(request)
    if not merchant:
        messages.error(request, '无权访问')
        return redirect('merchant_dashboard')

    item_ids = request.POST.getlist('item_ids')
    algorithm = request.POST.get('algorithm', 'maxrects')

    if not item_ids:
        messages.error(request, '请至少选择一个订单项进行拼版')
        return redirect('pending_layout_orders')

    items = OrderItem.objects.filter(
        id__in=item_ids,
        order__merchant=merchant,
        plate_file__isnull=False,
    ).exclude(plate_file='').select_related('order')

    if not items.exists():
        messages.error(request, '所选订单项无效或没有制版文件')
        return redirect('pending_layout_orders')

    first_item = items.first()
    product_name = first_item.product_name
    material = first_item.material
    thickness = first_item.thickness

    for item in items:
        if (item.product_name, item.material, item.thickness) != (product_name, material, thickness):
            messages.error(request, '请选择相同产品类型、材质和厚度的订单项进行拼版')
            return redirect('pending_layout_orders')
        if item.plate_batch and item.plate_batch.status == 'confirmed':
            messages.error(request, f'订单 {item.order.sn} 已确认拼版，无法重复拼版')
            return redirect('pending_layout_orders')

    try:
        from utils.plate_type_rules import get_spacing_mm
        from utils.plate_batch import build_rects_from_items, pack_rects_into_plates, persist_plate_batch

        spacing_mm = get_spacing_mm(thickness, thickness, material, material)
        rects = build_rects_from_items(items, spacing_mm=spacing_mm)
        if not rects:
            messages.error(request, '无法构建拼版矩形，请检查订单项尺寸')
            return redirect('pending_layout_orders')

        plates = pack_rects_into_plates(rects, algorithm=algorithm)
        if not plates:
            messages.error(request, '拼版失败，无法放置所有矩形')
            return redirect('pending_layout_orders')

        created_batches = []
        total_unplaced = 0
        for plate in plates:
            if not plate.get('placed'):
                continue
            total_unplaced += len(plate.get('unplaced', []))
            batch = persist_plate_batch(
                merchant, product_name, material, thickness, plate,
                algorithm=algorithm, use_plate_file=True,
            )
            if batch:
                created_batches.append(batch)

        if not created_batches:
            messages.error(request, '拼版失败，未能生成批次')
            return redirect('pending_layout_orders')

        usage = created_batches[0].usage_rate or 0
        msg = f'拼版成功！共 {len(created_batches)} 张版，首张利用率 {usage:.1f}%'
        if total_unplaced:
            msg += f'（{total_unplaced} 件未能放置，请减少数量或换更大板材）'
        messages.success(request, msg)
        return redirect('plate_batch_detail', batch_id=created_batches[0].id)

    except Exception:
        logger.exception('拼版批次创建失败')
        messages.error(request, '拼版失败，请稍后重试或联系客服')
        return redirect('pending_layout_orders')


@login_required
@merchant_required
def complaint_list(request):
    """
    商户投诉列表
    显示当前商户的所有客户投诉，含分类统计
    """
    try:
        merchant = request.user.managed_merchant if request.user.user_type == 'merchant_admin' else request.user.staff_profile.merchant
    except AttributeError:
        messages.error(request, '无权访问')
        return redirect('merchant_dashboard')

    # 获取该商户的所有订单的投诉
    complaints = OrderComplaint.objects.filter(
        order__merchant=merchant
    ).select_related('order', 'customer').order_by('-created_at')

    # 状态统计
    pending_count = complaints.filter(status='pending').count()
    processing_count = complaints.filter(status='processing').count()
    resolved_count = complaints.filter(status='resolved').count()
    rejected_count = complaints.filter(status='rejected').count()
    total_count = complaints.count()

    # 按投诉类型统计（只统计待处理和处理中的）
    from django.db.models import Count
    active_complaints = complaints.filter(status__in=['pending', 'processing'])
    type_stats = []
    type_counts = active_complaints.values('complaint_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # 类型分组统计
    file_count = 0
    prod_count = 0
    shipping_count = 0
    other_count = 0
    
    type_map = dict(OrderComplaint.COMPLAINT_TYPE_CHOICES)
    for tc in type_counts:
        ctype = tc['complaint_type'] or 'other'
        label = type_map.get(ctype, ctype)
        type_stats.append({'label': label, 'count': tc['count'], 'code': ctype})
        if ctype.startswith('file_'):
            file_count += tc['count']
        elif ctype.startswith('prod_'):
            prod_count += tc['count']
        elif ctype.startswith('shipping_'):
            shipping_count += tc['count']
        else:
            other_count += tc['count']

    # 近30天投诉趋势
    from django.utils import timezone
    from datetime import timedelta
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_complaints = complaints.filter(created_at__gte=thirty_days_ago)
    recent_count = recent_complaints.count()

    return render(request, 'merchant/complaint_list.html', {
        'complaints': complaints,
        'pending_count': pending_count,
        'processing_count': processing_count,
        'resolved_count': resolved_count,
        'rejected_count': rejected_count,
        'total_count': total_count,
        'recent_count': recent_count,
        'type_stats': type_stats,
        'file_count': file_count,
        'prod_count': prod_count,
        'shipping_count': shipping_count,
        'other_count': other_count,
    })


@login_required
@merchant_required
def complaint_detail(request, complaint_id):
    """
    商户投诉详情
    查看投诉详情并处理
    """
    try:
        merchant = request.user.managed_merchant if request.user.user_type == 'merchant_admin' else request.user.staff_profile.merchant
    except AttributeError:
        messages.error(request, '无权访问')
        return redirect('merchant_dashboard')

    complaint = get_object_or_404(
        OrderComplaint, id=complaint_id, order__merchant=merchant
    )

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'process_complaint':
            status = request.POST.get('status')
            merchant_remark = request.POST.get('merchant_remark', '').strip()

            if not merchant_remark:
                messages.error(request, '请填写处理备注')
                return redirect('merchant_complaint_detail', complaint_id=complaint.id)

            complaint.status = status
            complaint.merchant_remark = merchant_remark
            complaint.resolved_at = timezone.now()
            complaint.resolved_by = request.user
            complaint.save()

            messages.success(request, '投诉处理成功')
            return redirect('merchant_complaint_detail', complaint_id=complaint.id)

    return render(request, 'merchant/complaint_detail.html', {
        'complaint': complaint,
    })
